# 다운로드 폴더에 존재하는 파일을 insert 시키는 로직
# insert 시 데이터 중복에러가 나는 파일을 중복 제거 후 insert 한다.
# iaif_file에서 table_col에 part_id 위치 주의
# 주의사항 : 엑셀파일을 직접 다운로드 받아 수동 insert할 경우 한 번 저장한 후 실행해야 한다. 저장하지 않으면 open시 오류 발생
import time
import sys
import os
import openpyxl as pyxl
import cx_Oracle
import ctypes


from info_func import get_col_title, get_insert_text, element_check, is_univ_nm_check, is_univ_nm_space_check, \
    cell_extract, reg_d, reg_w, input_func, iaif_path_append
import infoClickYearly

os.environ["NLS_LANG"] = ".AL32UTF8"  # DB 케릭터셋과 일치시킴

wb, ws, wb_info, ws_info, v_cell_row, browser, bs_soup = "", "", "", "", 0, "", ""
dsn, conn = "", ""
result, insert_data, column_length = "", "", ""
v_year, v_table, univ_grp_type, part_id, v_mod = "", "", "", "", ""

part_list = ["1", "2", "3", "4", "5"]
part_name = {"1": "인문ㆍ사회계열", "2": "자연과학계열", "3": "공학계열", "4": "예ㆍ체능계열", "5": "의학계열"}

v_download_folder = r"V:\document\정보공시데이터\PYTHON\excel\down_file"
v_file_path = v_download_folder + "\\"  # 엑셀 다운로드 경로

print("*** IAIF5406S 교원 강의 담당 비율 학과별 : 읽기 전용(Y) ***")
print("!! 수동 다운로드 한 파일의 경우 한 번 열었다가 저장해야 데이터가 읽힙니다.")
# 파라미터
v_input_file_name = input("다운 파일명(기본확장자 xlsx): ")  # 파일명을 입력하면 해당 파일을 직접 insert한다.
v_input_file_name = v_input_file_name + ".xlsx"
# v_sheet_name = "Sheet"+input("시트번호[1,2,..] : ")
if v_input_file_name:
    v_input_univ_nm_index = input("대학명 위치 인덱스(default 5): ")
    if not v_input_univ_nm_index:
        v_input_univ_nm_index = 5
    v_input_univ_nm_index = int(v_input_univ_nm_index)

v_year = input("년도: ")
# v_dtYear = 2018  # 기준년도 : 공시에 따라 info_year 또는 info_year-1
v_table = input("공시테이블: ")
v_table = v_table.lower()
# 계열 "1": "인문ㆍ사회계열", "2": "자연과학계열", "3": "공학계열", "4": "예ㆍ체능계열", "5": "의학계열"
v_part_id = input("계열 코드(1~5 or Enter[해당없음]): ")
part_id = v_part_id
v_cols_order = input('칼럼 목록 순서[2, 3 or Enter] : ')  # 항목 칼럼이 상이할 시 col, col2, ...에서 지정하는 순서
v_read_only = input_func("읽기 전용 여부(Y|N)", "str", ["Y", "N"])
v_excel_save_yn = input_func("공시관리 엑셀 저장 여부(Y|N)", "str", ["Y", "N"])

v_univgrp1 = ""  # 태그의 option value - 01[전문대학], 02[대학], 03[대학원]
if v_table[4:5] == "5":
    v_univgrp1 = "대학"  # "02"
elif v_table[4:5] == "7":
    v_univgrp1 = "전문대학"  # "01"
else:
    v_univgrp1 = "대학원"  # "03"

v_sheet_name = "학과별_" + v_univgrp1

# 공시 파일 내 변수 가져오기
iaif_path_append()
v_mod = __import__(v_table)

v_iaif_name = v_mod.iaif_name
v_total_col = v_mod.total_col

try:
    v_dtyear_idx = v_mod.dtyear_idx
    v_dtyear = str(int(v_year) - int(v_mod.dtyear_num))
except AttributeError:
    v_dtyear = str(int(v_year) - int(v_mod.dtyear_num))
    v_dtyear_idx = -1  # 공시를 조회한 항목에서 넣어야 할 때, 그 외는 dtyear_num값 만큼 평가년도에서 감소


def insert_logic(m_mod, m_part_id):  # 테이블 insert 처리
    print(insert_data[:3])
    if len(insert_data):
        # show_data(insert_data)
        v_table_cols = ""
        if v_cols_order == "":
            v_table_cols = "(" + ",".join(v_mod.table_cols) + ")" if v_mod.table_cols else ""
        elif v_cols_order == "2":
            v_table_cols = "(" + ",".join(v_mod.table_cols2) + ")" if v_mod.table_cols2 else ""
        elif v_cols_order == "3":
            v_table_cols = "(" + ",".join(v_mod.table_cols3) + ")" if v_mod.table_cols3 else ""

        insert_sql = "INSERT INTO " + v_table + v_table_cols + " VALUES ('" + str(v_year) + "'," + \
                     get_insert_text(column_length) + m_mod.insert_last_col + ")"
        cursor_ins = conn.cursor()
        try:
            # insert_data = unicode(insert_data, "euc-kr").encode("utf-8")
            cursor_ins.executemany(insert_sql, insert_data)
            cursor_ins.close()
            conn.commit()
            print(v_table + " " + (part_name[m_part_id] if len(part_id) else "") + " insert success!")
            if v_excel_save_yn == "Y":
                ws.cell(v_cell_row, v_year_index_in_file + 1).value = "O"  # 엑셀 칼럼은 1부터 시작하므로 년도 추출 변수에 1을 더함
                wb.save("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")

            # 엑셀 파일 삭제 여부
            returnVal = ctypes.windll.user32.MessageBoxW(0, str(v_year) + "다운로드 받은 엑셀 파일을 삭제할까요?", "", 4100)
            if returnVal == 6:  # 예:6, 아니오:7
                os.remove(v_file_path + v_input_file_name)
                print(v_input_file_name, " 파일이 삭제되었습니다.")
            elif returnVal == 7:
                pass

            # 엑셀 insert 후 오류 확인 등을 위해 파일은 남겨둠
            # if os.path.isfile(v_file_path + v_input_file_name):
            #     os.remove(v_file_path + v_input_file_name)
        except cx_Oracle.DatabaseError as e:
            error, = e.args
            print(insert_sql)
            print(error.code)
            print(error.message)
            print(error.context)
            cursor_ins.close()
            conn.rollback()
            conn.close()
            if v_excel_save_yn == "Y":
                wb.close()
            sys.exit()
    else:
        print("엑셀 파일에 데이터가 없거나 재저장이 필요한 파일입니다.")


def insert_msg( m_mod, m_part_id):  # 공시와 테이블 비교에 따른 구분
    try:
        if v_mod.delete_except:
            print("해당 항목은 Delete 제외 공시입니다.")
            pass
    except AttributeError:
        # 기데이터 삭제
        cursor_del = conn.cursor()
        part_sql = "AND PART_ID = " + m_part_id if len(part_id) else ""
        cursor_del.execute("DELETE FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) + "'" + part_sql)
        time.sleep(2)
        cursor_del.close()
    time.sleep(0.33)
    insert_logic(m_mod, m_part_id)


def main_logic():
    # 대학원 취업 현황의 전문대학원과 특수대학원은 따로 공시되지만 같은 공시 테이블을 사용하므로
    # 뒤에 오는 특수대학원에 공시 테이블명을 두어 다른 파일명이지만 같은 테이블에 insert 되도록 처리

    # 대학명이 존재하는 칼럼 index 할당 : 일부 공시는 학교종류가 병합될 수 도 있음
    # 보통 데이터는 대학명부터 잘라서 집어넣는데, 대학원(신입생 충원 현황)의 경우 대학원종류 칼럼부터 insert하므로 이를 처리
    try:
        univ_nm_idx = v_mod.main_data_start_index
    except AttributeError:
        univ_nm_idx = v_input_univ_nm_index

    global insert_data
    insert_data = []  # table insert data
    global column_length
    column_length = 0  # insert 행의 컬럼 길이

    # 다운받은 파일을 열어서 데이터 추출
    global wb_info, ws_info
    v_read_only_chk = True if v_read_only == "Y" else False
    wb_info = pyxl.load_workbook(filename=v_file_path + v_input_file_name, read_only=v_read_only_chk)
    print('file opened')
    for sheet_name in wb_info.sheetnames:
        print("<<", sheet_name, "start >>")
        ws_info = wb_info[sheet_name]  # or wb.active : 활성화된 시트

        # 데이터 시작 로우 번호 찾기
        v_compare_row_num = 0
        global v_dtyear
        for idx, row in enumerate(list(ws_info.rows)):
            if type(row[0].value) == int:
                if row[0].value == v_dtyear or row[0].value == v_year:
                    v_compare_row_num = idx
                    break
            elif row[0].value is not None and row[0].value.isnumeric():
                if row[0].value == str(v_dtyear) or row[0].value == str(v_year):
                    v_compare_row_num = idx
                    break

        v_compare_row_num = v_mod.excel_data_start_row

        print('데이터 시작 행 : ', v_compare_row_num)
        merged_cell_list = []  # 데이터 중 병합이 있는 범위를 담는 list

        # 데이터 중 칼럼의 병합은 불가. 로우의 병합만 존재.
        if v_read_only == "N":
            for m_row in ws_info.merged_cells.ranges:
                v_cell_range = cell_extract(m_row)
                if v_cell_range[1] >= v_compare_row_num+1:  # 셀은 1부터 시작하므로 +1
                    merged_cell_list.append(v_cell_range)  # [칼럼 알파벳, 병합시작row, 병합종료row]

        v_data_chk_list = []  # 중복 체크를 위한 키 묶음 텍스트(v_data_key_txt) 저장 list
        v_data_dup_list = {}  # 중복되는 데이터 담는 dictionary

        ws_row_list = list(ws_info.rows)

        sust_index = None
        for idx, row in enumerate(ws_row_list[3]):
            if row.value is not None and "학과" in row.value.replace(" ", ""):
                sust_index = idx
                break

        if len(ws_row_list[len(ws_row_list) - 1]) == 0:  # tuple ()
            v_last_row_except_num = 1
        else:
            v_last_row_except_num = 1 if ws_row_list[len(ws_row_list) - 1][0].value is None else 0

        print('엑셀 데이터 수 :', len(ws_row_list))
        for idx, row in enumerate(ws_row_list[v_compare_row_num:len(ws_row_list) - v_last_row_except_num]):
            item_list = []
            v_data_key_txt = ""  # 키 묶음 텍스트 변수

            # print('인덱스 : ', idx)

            # 행 제외 데이터에 포함되면 해당 행 부분은 insert하지 않고 다음 반복으로 이동
            # iaif 파일에 delete_row_compare_col_index 변수 선언
            # 빈 값이거나 지정된 칼럼 명일 경우 제외, insert하는 데이터는 지정 칼럼에 null(None)이 없어야 한다.
            # iaif5406s
            try:
                if (str(row[v_mod.delete_row_compare_col_index[0]].value) in v_mod.delete_row_compare_col_index[1] or
                        row[v_mod.delete_row_compare_col_index[0]].value is None):
                    continue
            except AttributeError:
                pass
            except IndexError:  # 행의 데이터 수가 부족한 행은 일반적인 데이터 행이 아니므로 제외
                continue

            # 기준년도 칼럼 추출. (재학생 충원 현황: iaif5108s_2013)
            if v_table == "iaif5108s_2013":
                v_dtyear = row[v_dtyear_idx].value if v_dtyear_idx > -1 else v_dtyear
                item_list.append(v_dtyear)
            else:
                pass

            for d_idx, d_value in enumerate(list(row)[univ_nm_idx:]):
                global row_value
                row_value = str(d_value.value)

                if merged_cell_list:
                    if not row_value:
                        addr_w = reg_w.search(d_value.coordinate).group()
                        addr_d = int(reg_d.search(d_value.coordinate).group())
                        for m_arr in merged_cell_list:
                            if addr_w == m_arr[0] and m_arr[1] <= addr_d <= m_arr[2]:  # 병합된 셀
                                row_value = insert_data[len(insert_data)-1][d_idx]  # 이전 행의 해당 인덱스의 요소 할당
                                break
                else:
                    if d_idx+univ_nm_idx in v_mod.data_key_index:
                        # 데이터 행에 대한 시작 열을 지정해서 가져오기 때문에 그 수치만큼 더하여야 한다.(univ_nm_idx)
                        if row_value in ["None", ""]:
                            row_value = insert_data[len(insert_data) - 1][d_idx+univ_nm_idx]  # 이전 행의 해당 인덱스의 요소 할당

                if is_univ_nm_check(row_value):
                    row_value = row_value.replace(" _", "_")
                elif v_year == 2016 and v_table in ("iaif5200_13", "iaif6200"):
                    # 졸업생 취업률의 취업률 칼럼은 16년 이전에는 '-'로 표시되었기 때문에 ''로 수정
                    row_value = "" if row_value == "-" else row_value.replace(",", "")
                else:
                    if row_value == "None":
                        row_value = ""
                    else:
                        if sust_index is None or d_idx != sust_index - univ_nm_idx:
                            # 수치 데이터의 ','(comma) 표시 제거
                            row_value = row_value.replace(",", "")  # .replace("∙", "ㆍ")

                # v_mod.data_key_index는 IAIF 파일 변수인데 이는 엑셀 다운로드 데이터 기준 인덱스이기 때문에
                # 현재 루프의 d_idx에서 시작하는 열로 지정한 univ_nm_idx만큼 더해야 한다.
                if (d_idx + univ_nm_idx) in v_mod.data_key_index:
                    # row_value가 NoneType일 경우 오류 발생 : 이전 로직에서 None을 반환하였기 때문
                    v_data_key_txt += row_value if v_data_key_txt == "" else "/" + row_value

                item_list.append(row_value)

            if v_data_key_txt in v_data_chk_list:
                if v_data_key_txt in v_data_dup_list.keys():
                    v_data_dup_list[v_data_key_txt] += 1
                else:
                    v_data_dup_list[v_data_key_txt] = 2
            else:
                v_data_chk_list.append(v_data_key_txt)

            # 일부 공시는 테이블이 분리되어 있기 때문에 나눠서 추출한다. (국내외 연구실적, 저역서 실적)
            v_div_list_module, v_data_div_list = [], []
            try:
                if v_mod.data_div_rng:
                    # v_cols_order의 순서에 따라 data_div_rng의 번호를 할당
                    if v_cols_order == "":
                        v_div_list_module = v_mod.data_div_rng
                    elif v_cols_order == "2":
                        v_div_list_module = v_mod.data_div_rng2
                    elif v_cols_order == "3":
                        v_div_list_module = v_mod.data_div_rng3

                for idx, value in enumerate(v_div_list_module):
                    if idx == 0:
                        v_data_div_list += item_list[:value - univ_nm_idx]
                    else:
                        v_data_div_list += item_list[value[0] - univ_nm_idx:value[1] - univ_nm_idx]
            except AttributeError:
                v_data_div_list = None
                pass

            if len(item_list) > 0:
                # print(item_list)
                v_dtyear_element = [v_dtyear] if v_table != "iaif5108s_2013" else []
                insert_data.append((v_dtyear_element + item_list if not v_data_div_list else v_data_div_list)
                                   + (list(str(part_id)) if len(part_id) else []))
            else:
                continue

            # print(item_list)
            if len(insert_data) == 1:
                # print(insert_data[0])
                column_length = len(insert_data[0])
            # print(insert_data[len(insert_data)-1])

        # v_mod(공시 항목 파일)에 data_dup_except 변수가 True이면 중복 Pass
        # 해당 변수가 없고, 중복 리스트에 값이 존재하면 중단.
        try:
            if v_mod.data_dup_except:
                pass
        except AttributeError:
            if v_data_dup_list:
                print("중복되는 데이터가 존재합니다.")
                print(v_data_dup_list)
                conn.close()
                wb.close()
                sys.exit()
            else:
                pass

    wb_info.close()
    # for elmt in insert_data[:10]:
    #     print(elmt)
    insert_msg(v_mod, part_id)
# end-function


if infoClickYearly.v_total_insert_yn == "Y":

    # oracle 연동
    dsn = cx_Oracle.makedsn("61.81.234.137", 1521, "COGDW")
    conn = cx_Oracle.connect("dusd", "dusd$#@!", dsn)

    if v_excel_save_yn == "Y":
        # 정보공시 재업데이트 공시 목록 불러오기
        wb = pyxl.load_workbook("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")
        ws = wb[v_sheet_name]  # or wb.active : 활성화된 시트
        v_year_index_in_file = 2

        # infoClickYearly 파일의 지정 년도가 재업데이트 파일에서의 칼럼의 년도에서의 위치 추출
        for idx, col in enumerate(list(ws.rows)[0]):
            if str(col.value) == str(infoClickYearly.v_down_year):
                v_year_index_in_file = idx
                break

        for r in list(ws.rows)[1:]:
            # insert_chk = r[v_year_index_in_file].value
            v_file_iaif_name = r[0].value.lower()
            global v_file_part_id, v_part_check
            v_part_check = True
            if len(part_id):
                v_file_part_id = str(r[1].value)
                v_part_check = str(v_file_part_id) == v_part_id
            # 엑셀파일에서 해당 년도 값이 비었으면 항목 추출 실행
            if v_file_iaif_name == v_table and v_part_check and r[v_year_index_in_file].value is None:
                v_cell_row = r[0].row
                if len(part_id):
                    print("공시 항목 : " + v_file_iaif_name + ", 계열 : " + part_name[str(v_file_part_id)])
                else:
                    print("공시 항목 : " + v_file_iaif_name)
                main_logic()
    else:
        if len(part_id):
            print("공시 항목 : " + v_table + ", 계열 : " + part_name[part_id])
        else:
            print("공시 항목 : " + v_table)
        main_logic()

conn.close()
try:
    wb.close()
except AttributeError:
    sys.exit()
