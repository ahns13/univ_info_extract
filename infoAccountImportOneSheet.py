# 재정알리미 다운로드 후 엑셀자료를 추출하여 import하는 파일 - 하나의 시트에 다운로드한 파일 버전
import time
import sys
import os
import openpyxl as pyxl
import xlrd
import cx_Oracle
import tkinter as tk
import tkinter.ttk as ttk
from info_func import get_insert_text, input_func, iaif_path_append

os.environ["NLS_LANG"] = ".AL32UTF8"  # DB 케릭터셋과 일치시킴

v_start_row = 2
v_univ_cols = [1, 2, 3, 4, 6]  # F열은 병합처리되어 제외


def box_close(event=None):
    global v_file_name
    v_file_name = comboSel.get()
    combo.destroy()


def go_exit():
    sys.exit()


univ_grp = {"A": "대학", "B": "전문대"}
v_input_univ_grp = ""
dic_file_type = {
    "1": ["V:\\document\\정보공시데이터\\", "\\사립대학회계정보시스템 추출파일\\자금예산서_교비회계(예산)\\"],
    "2": ["V:\\document\\정보공시데이터\\", "\\사립대학회계정보시스템 추출파일\\자금계산서_교비회계(결산)\\"],
    "3": ["V:\\document\\정보공시데이터\\", "\\사립대학회계정보시스템 추출파일\\산학협력단회계_현금흐름표_지출\\"]
}

v_file_type = input_func("1: 교비 자금예산서(예산), 2: 교비 자금계산서(결산), 3:산학협력단 자금계산서(지출) : ", "str", ["1", "2", "3"])
input_univ_grp = input_func("A: 대학, B: 전문대학 : ", "str", ["A", "B"])
v_input_univ_grp = univ_grp[input_univ_grp]
v_file_path = dic_file_type[v_file_type][0] + v_input_univ_grp + dic_file_type[v_file_type][1]

v_year = input("폴더 년도 : ")

print("창에 나타난 파일 목록 중 하나를 선택하세요.")
combo = tk.Tk()
combo.geometry("450x120")
combo.title("파일 선택")
label1 = tk.Label(combo, text="파일명: ", width=12).grid(row=1, column=0)
comboSel = ttk.Combobox(combo, values=os.listdir(v_file_path + str(v_year)), width=35)
comboSel.grid(row=1, column=1)
comboSel.current(0)
combo.bind("<Return>", box_close)
btn = tk.Button(combo, text="추출 실행", command=box_close).grid(row=3, column=1)
btn2 = tk.Button(combo, text="종료", command=go_exit).grid(row=4, column=1)
combo.lift()
combo.focus_force()
combo.mainloop()

# input("파일명(확장자포함) : ")
print(v_file_name)
v_table = input("테이블명[iaif5538,iaif5539,iaif5552,iaif7538,iaif7552] : ")
v_table = v_table.lower()

v_finan_num = input_func("1.수입, 2.지출(숫자 입력) : ", "str", ["1", "2"])
v_finan_gbn = {"1": "수입", "2": "지출"}
v_finan_name = v_finan_gbn[v_finan_num]

# 공시 파일 내 변수 가져오기
iaif_path_append()
v_mod = __import__(v_table)

# v_sheet_double_chk = v_mod.sheet_double_chk
v_dt_year_num = v_mod.dt_year_num
# v_finan_item_list = v_mod.finan_gbn[v_finan_name]

v_file_full_path = v_file_path + str(v_year) + "\\" + v_file_name

# oracle 연동
dsn = cx_Oracle.makedsn("61.81.234.137", 1521, "COGDW")
conn = cx_Oracle.connect("dusd", "dusd$#@!", dsn)

v_table_cols = "(" + ",".join(v_mod.table_cols) + ")" if v_mod.table_cols else ""

insert_sql = "INSERT INTO " + v_table + v_table_cols + " VALUES (" + get_insert_text(len(v_mod.table_cols)-2) + v_mod.insert_last_col + ")"
cursor = conn.cursor()

wb = xlrd.open_workbook(v_file_full_path)

v_univ_name = []
insert_data = []
v_univ_data = {
    "univ_data1": [],
    "univ_data2": [],
    "univ_data3": [],
    "univ_data4": [],
    "univ_data5": []
}

v_sheet_start_chk = False

print("추출 start")
ws = wb.sheet_by_index(0)  # or wb.active : 활성화된 시트

v_univ_name = []
insert_data = []
v_univ_data = {
    "univ_data1": [],
    "univ_data2": [],
    "univ_data3": [],
    "univ_data4": [],
    "univ_data5": []
}


def univData_insert(p_univ_data):
    global insert_data
    for key in p_univ_data.keys():
        insert_data = insert_data + p_univ_data[key]


lvl_name_dic = {}
v_item_list = []

for r_idx in range(0, ws.nrows):
    if ws.cell(r_idx, 0).value == "항목":
        univData_insert(v_univ_data)

        v_univ_data["univ_data1"] = []
        v_univ_data["univ_data2"] = []
        v_univ_data["univ_data3"] = []
        v_univ_data["univ_data4"] = []
        v_univ_data["univ_data5"] = []
        v_univ_name = []
        v_item_list = []

        for c_no in v_univ_cols:
            try:
                r_univ_name = ws.cell(r_idx, c_no).value
                if len(r_univ_name) > 3:
                    if "[산단]" in r_univ_name:
                        r_univ_name = r_univ_name.replace("[산단]", "")
                    print(r_univ_name)
                    v_univ_name.append(r_univ_name + "_본교")
                else:
                    continue
            except:
                break
    elif ws.cell(r_idx, 0).value in ["None", "", "교비회계 자금계산서( 수입 )", "교비회계 자금계산서( 지출 )", "산학협력단회계 자금계산서( 지출 )"]:
        continue
    else:
        v_item = ws.cell(r_idx, 0).value
        v_mod_item = v_item.lstrip()
        v_item_list.append(v_mod_item)
        v_size_item = len(v_item)
        v_size_mod_item = len(v_mod_item)
        v_level = int((v_size_item - v_size_mod_item)/2)  # 2개 공백이 한개 레벨임
        v_level_item_name = str(v_level) + ". " + v_mod_item

        lvl_name_dic[v_level] = v_mod_item

        if v_file_type == "3":  # 산단 자금계산서
            for c_idx, c_univ in enumerate(v_univ_name):
                v_univ_data["univ_data" + str(c_idx+1)].append([
                    v_year, str(int(v_year) - v_dt_year_num), v_finan_name, c_univ,
                    v_mod_item, v_level, v_level_item_name,
                    "최상위지표" if v_level == 1 else str(v_level-1) + ". " + lvl_name_dic[v_level-1],
                    "최상위지표" if v_level in [1, 2] else str(v_level-2) + ". " + lvl_name_dic[v_level-2],
                    len(v_item_list),
                    None if ws.cell(r_idx, v_univ_cols[c_idx]).value == "null" else ws.cell(r_idx, v_univ_cols[c_idx]).value
                ])
        else:
            for c_idx, c_univ in enumerate(v_univ_name):
                v_univ_data["univ_data" + str(c_idx+1)].append([
                    v_year, str(int(v_year) - v_dt_year_num), v_finan_name, c_univ,
                    v_mod_item, v_level, v_level_item_name,
                    "최상위지표" if v_level == 1 else str(v_level-1) + ". " + lvl_name_dic[v_level-1],
                    len(v_item_list),
                    None if ws.cell(r_idx, v_univ_cols[c_idx]).value == "null" else ws.cell(r_idx, v_univ_cols[c_idx]).value
                ])
            # for c_idx, c_univ in enumerate(v_univ_name):
            #     v_univ_data["univ_data" + str(c_idx+1)].append([
            #         v_year, str(int(v_year) - v_dt_year_num), v_finan_name, c_univ,
            #         v_mod_item, v_level, v_level_item_name, v_finan_item_list[v_level_item_name][2],
            #         None if ws.cell(r_idx, v_univ_cols[c_idx]).value == "null" else ws.cell(r_idx, v_univ_cols[c_idx]).value
            #     ])

    if r_idx == ws.nrows-1:
        univData_insert(v_univ_data)

try:
    for u_data in insert_data[:3]:
        print(u_data)
    cursor.executemany(insert_sql, insert_data)
    conn.commit()
    print("insert success!")
except cx_Oracle.DatabaseError as e:
    error, = e.args
    print(insert_sql)
    print(error.code)
    print(error.message)
    print(error.context)
    print("error!")
    cursor.close()
    conn.rollback()
    conn.close()
    wb.release_resources()
    sys.exit()


conn.close()
wb.release_resources()
