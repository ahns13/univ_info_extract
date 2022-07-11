# 다운로드 폴더에 존재하는 파일을 insert 시키는 로직
# insert 시 데이터 중복에러가 나는 파일을 중복 제거 후 insert 한다.
# 주의사항 : 엑셀파일을 직접 다운로드 받아 수동 insert할 경우 한 번 저장한 후 실행해야 한다. 저장하지 않으면 open시 오류 발생
import time
import sys
import os
import openpyxl as pyxl
import keyboard
import pyperclip
import cx_Oracle


from info_func import get_col_title, get_insert_text, element_check, is_univ_nm_check, is_univ_nm_space_check, \
    cell_extract, reg_d, reg_w,input_func, iaif_path_append

os.environ["NLS_LANG"] = ".AL32UTF8"  # DB 케릭터셋과 일치시킴

wb, ws, wb_info, ws_info, v_cell_row, browser, bs_soup = "", "", "", "", 0, "", ""
dsn, conn = "", ""
result, insert_data, column_length = "", "", ""
v_year, v_table, univ_grp_type, v_mod = "", "", "", ""

v_download_folder = r"V:\document\정보공시데이터\PYTHON\excel\down_file"
v_file_path = v_download_folder + "\\"  # 엑셀 다운로드 경로

# 파라미터
v_input_file_name = input("다운 파일명(기본확장자 xlsx): ")  # 파일명을 입력하면 해당 파일을 직접 insert한다.
v_input_file_name = v_input_file_name + ".xlsx"
if v_input_file_name:
    v_input_univ_nm_index = input("대학명 위치 인덱스(default 5): ")

v_year = input("년도: ")
# v_dtYear = 2018  # 기준년도 : 공시에 따라 info_year 또는 info_year-1
v_table = input("공시테이블: ")
v_table = v_table.lower()

v_cols_order = input('칼럼 목록 순서[2, 3 or Enter] : ')  # 항목 칼럼이 상이할 시 col, col2, ...에서 지정하는 순서
# v_row_start_num = input('데이터 행 시작 index(0~)[파일 참조 시 Enter] : ')  # 데이터 행이 시작하는 인덱스 0부터 기입
v_save_exec_yn = input_func("파일 재저장 실행 여부(Y|N) : ", "str", ["Y", "N"])
# 다운로드 받은 파일을 저장하고 열어야 cumstomWidth 오류가 없음
v_excel_save_yn = input_func("공시관리 엑셀 저장 여부(Y|N) : ", "str", ["Y", "N"])

v_sust_yn = input_func("학과 공시 여부(Y|N) : ", "str", ["Y", "N"])

v_univgrp1 = ""  # 태그의 option value - 01[전문대학], 02[대학], 03[대학원]
if v_table[4:5] == "5":
    v_univgrp1 = "대학"  # "02"
elif v_table[4:5] == "7":
    v_univgrp1 = "전문대학"  # "01"
else:
    v_univgrp1 = "대학원"  # "03"

v_sheet_name = ""
if v_sust_yn == "Y":
    v_sheet_name = "학과별_" + v_univgrp1
else:
    v_sheet_name = v_univgrp1

# 공시 파일 내 변수 가져오기
iaif_path_append()
v_mod = __import__(v_table)

v_iaif_name = v_mod.iaif_name
v_total_col = v_mod.total_col

try:
    v_dtyear_idx = v_mod.dtyear_idx
except AttributeError:
    v_dtyear = int(v_year) - int(v_mod.dtyear_num)
    v_dtyear_idx = -1  # 공시를 조회한 항목에서 넣어야 할 때, 그 외는 dtyear_num값 만큼 평가년도에서 감소

if not v_input_univ_nm_index:
    try:
        v_input_univ_nm_index = v_mod.column_start_col
    except AttributeError:
        v_input_univ_nm_index = 5
v_input_univ_nm_index = int(v_input_univ_nm_index)


def insert_logic(m_mod):  # 테이블 insert 처리
    # print(insert_data[:3])
    for r in insert_data[:10]:
        print(r)
    v_table_cols = ""
    if v_cols_order == "":
        v_table_cols = """(""" + ",".join(v_mod.table_cols) + """)""" if v_mod.table_cols else ""
    elif v_cols_order == "2":
        v_table_cols = """(""" + ",".join(v_mod.table_cols2) + """)""" if v_mod.table_cols2 else ""
    elif v_cols_order == "3":
        v_table_cols = """(""" + ",".join(v_mod.table_cols3) + """)""" if v_mod.table_cols3 else ""

    insert_sql = "INSERT INTO " + v_table + v_table_cols + " VALUES ('" + str(v_year) + "','" + str(v_dtyear) + "'," + \
                 get_insert_text(column_length) + m_mod.insert_last_col + ")"
    cursor_ins = conn.cursor()
    try:
        # insert_data = unicode(insert_data, "euc-kr").encode("utf-8")
        cursor_ins.executemany(insert_sql, insert_data)
        cursor_ins.close()
        conn.commit()
        print(v_table + " " + " update success!")

        if v_excel_save_yn == "Y":
            if v_excel_insert_check:
                ws.cell(v_cell_row, v_year_index_in_file + 1).value = "O"  # 엑셀 칼럼은 1부터 시작하므로 년도 추출 변수에 1을 더함
            wb.save("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")

        if os.path.isfile(v_file_path + v_input_file_name):
            os.remove(v_file_path + v_input_file_name)
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(insert_sql)
        print(error.code)
        print(error.message)
        print(error.context)
        cursor_ins.close()
        conn.rollback()
        conn.close()
        if v_excel_save_yn == "Y":
            wb.close()
        sys.exit()


def insert_msg(m_mod):  # 공시와 테이블 비교에 따른 구분
    try:
        if v_mod.delete_except:
            print("해당 항목은 Delete 제외 공시입니다.")
            pass
    except AttributeError:
        # 기데이터 삭제
        cursor_del = conn.cursor()
        cursor_del.execute("DELETE FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) + "'")
    time.sleep(2)
    cursor_del.close()

    time.sleep(0.33)
    insert_logic(v_mod)


def main_logic():
    # 대학원 취업 현황의 전문대학원과 특수대학원은 따로 공시되지만 같은 공시 테이블을 사용하므로
    # 뒤에 오는 특수대학원에 공시 테이블명을 두어 다른 파일명이지만 같은 테이블에 insert 되도록 처리

    global insert_data
    insert_data = []  # table insert data
    global column_length
    column_length = 0  # insert 행의 컬럼 길이

    # 다운로드 한 파일을 한 번 열고 저장하는 과정을 거쳐야 cumstomWidth오류가 발생하지 않음
    if v_save_exec_yn == "Y":
        time.sleep(1.5)
        os.startfile(v_file_path + v_input_file_name)
        time.sleep(4)
        v_check = False
        v_count = 0
        while not v_check:
            keyboard.press_and_release("control+c")
            time.sleep(0.33)
            cell_text = pyperclip.paste()
            if v_year in cell_text or "info" in cell_text:
                time.sleep(0.8)
                keyboard.press_and_release("control+s")
                time.sleep(0.5)
                keyboard.press_and_release("alt+f+x")
                time.sleep(1)
                v_check = True

            v_count += 1
            if v_count > 120:
                print("저장된 파일을 실행시키는데 실패하였습니다.")
                conn.close()
                wb.close()
                browser.quit()
                sys.exit()
            time.sleep(0.33)

    # 다운받은 파일을 열어서 데이터 추출
    global wb_info, ws_info
    wb_info = pyxl.load_workbook(v_file_path + v_input_file_name)
    ws_info = wb_info.active  # or wb.active : 활성화된 시트

    # 데이터 시작 로우 번호 찾기
    v_row_start_num = 0
    for idx, row in enumerate(list(ws_info.rows)):
        if type(row[0].value) == int:
            if row[0].value == v_dtyear or row[0].value == v_year:
                v_row_start_num = idx
                break
        elif row[0].value is not None and row[0].value.isnumeric():
            if row[0].value == str(v_dtyear) or row[0].value == str(v_year):
                v_row_start_num = idx
                break

    v_compare_row_num = v_row_start_num if v_row_start_num else v_mod.excel_data_start_row
    print('데이터 시작 행 : ', v_compare_row_num)

    merged_cell_list = []  # 데이터 중 병합이 있는 범위를 담는 list
    # 데이터 중 칼럼의 병합은 불가. 로우의 병합만 존재.
    for m_row in ws_info.merged_cells.ranges:
        v_cell_range = cell_extract(m_row)
        if v_cell_range[1] >= v_compare_row_num+1:  # 셀은 1부터 시작하므로 +1
            merged_cell_list.append(v_cell_range)  # [칼럼 알파벳, 병합시작row, 병합종료row]

    ws_row_list = list(ws_info.rows)
    v_last_row_except_num = 1 if ws_row_list[len(ws_row_list) - 1][0].value is None else 0
    print('엑셀 데이터 수 :', len(ws_row_list))
    for idx, row in enumerate(ws_row_list[v_compare_row_num:len(ws_row_list) - v_last_row_except_num]):
        item_list = []

        if v_table == "iaif5354":  # 학비감면 준수 여부
            if "30%" in row[2].value:  # 30%규정준수금액의 학부+대학원 금액
                insert_data[len(insert_data)-1].append(str(row[v_mod.column_insert_data[1]].value))
                continue

        # iaif5522_21, iaif5529_17
        try:
            if (str(row[v_mod.delete_row_compare_col_index[0]].value) in v_mod.delete_row_compare_col_index[1] or
                    row[v_mod.delete_row_compare_col_index[0]].value is None):
                continue
        except AttributeError:
            pass
        except IndexError:  # 행의 데이터 수가 부족한 행은 일반적인 데이터 행이 아니므로 제외
            continue

        # 불필요한 열 병합으로 인한 빈 칼럼에 해당하는 데이터 제외하기
        # iaif 파일에 delete_col_index_list 변수 선언
        try:
            row = [el for no, el in enumerate(row) if no not in v_mod.delete_col_index_list]
        except AttributeError:
            pass

        for d_idx, d_value in enumerate(list(row)[v_input_univ_nm_index:]):
            global row_value
            row_value = str(d_value.value)

            if v_table == "iaif5354":
                if (d_idx+v_mod.column_start_col) in v_mod.column_insert_data[0]:  # 10% insert 칼럼
                    pass
                else:
                    continue

            if merged_cell_list and idx > 0: # 데이터의 맨 첫 행은 빈 값이어도 병합 처리 로직에서 제외(이전 행이 없기 때문)
                if row_value is None or row_value == "" or row_value == "None":
                    addr_w = reg_w.search(d_value.coordinate).group()
                    addr_d = int(reg_d.search(d_value.coordinate).group())
                    for m_arr in merged_cell_list:
                        if addr_w == m_arr[0] and m_arr[1] <= addr_d <= m_arr[2]:  # 병합된 셀
                            row_value = insert_data[len(insert_data)-1][d_idx]  # 이전 행의 해당 인덱스의 요소 할당
                            break

            if is_univ_nm_check(row_value):
                row_value = row_value.replace(" _", "_")
            elif v_year == 2016 and v_table in ("iaif5200_13", "iaif6200"):
                # 졸업생 취업률의 취업률 칼럼은 16년 이전에는 '-'로 표시되었기 때문에 ''로 수정
                row_value = "" if row_value == "-" else row_value.replace(",", "")
            elif v_table == "iaif5541":
                row_value = row_value.replace("-", "0")
            else:
                if row_value == "None":
                    row_value = ""
                else:
                    # 수치 데이터의 ','(comma) 표시 제거
                    row_value = row_value.replace(",", "")  # .replace("∙", "ㆍ")

            item_list.append(row_value)

        # 대학명이 존재하는 칼럼 index 할당 : 일부 공시는 학교종류가 병합될 수 도 있음
        # 보통 데이터는 대학명부터 잘라서 집어넣는데, 대학원(신입생 충원 현황)의 경우 대학원종류 칼럼부터 insert하므로 이를 처리
        try:
            univ_nm_idx = v_mod.main_data_start_index
        except AttributeError:
            univ_nm_idx = v_input_univ_nm_index

        # 일부 공시는 테이블이 분리되어 있기 때문에 나눠서 추출한다. (국내외 연구실적, 저역서 실적)
        v_div_list_module, v_data_div_list = [], []
        try:
            if v_mod.data_div_rng:
                # v_cols_order의 순서에 따라 data_div_rng의 번호를 할당
                if v_cols_order == "":
                    v_div_list_module = v_mod.data_div_rng
                elif v_cols_order == "2":
                    v_div_list_module = v_mod.data_div_rng2
                elif v_cols_order == "3":
                    v_div_list_module = v_mod.data_div_rng3

            for m_idx, value in enumerate(v_div_list_module):
                v_data_div_list += item_list if m_idx == 0 else item_list[value[0]:value[1]]
        except AttributeError:
            v_data_div_list = None
            pass

        if v_excel_save_yn == "Y":
            wb_info.close()
        insert_data.append(item_list if not v_data_div_list else v_data_div_list)

    column_length = len(insert_data[len(insert_data)-1])

    # for elmt in insert_data[:10]:
    #     print(elmt)
    insert_msg(v_mod)
# end-function


# oracle 연동
dsn = cx_Oracle.makedsn("61.81.234.137", 1521, "COGDW")
conn = cx_Oracle.connect("dusd", "dusd$#@!", dsn)

if v_excel_save_yn == "Y":
    # 정보공시 재업데이트 공시 목록 불러오기
    wb = pyxl.load_workbook("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")
    ws = wb[v_sheet_name]  # or wb.active : 활성화된 시트
    v_year_index_in_file = 2

    # infoClickYearly 파일의 지정 년도가 재업데이트 파일에서의 칼럼의 년도에서의 위치 추출
    for idx, col in enumerate(list(ws.rows)[0]):
        if str(col.value) == str(v_year):
            v_year_index_in_file = idx
            break

if v_excel_save_yn == "Y":
    for r in list(ws.rows)[1:]:
        insert_chk = r[v_year_index_in_file].value
        v_file_iaif_name = r[0].value.lower()

        global v_excel_insert_check
        if v_file_iaif_name in ("iaif5506s", "iaif7506s"):
            insert_chk = False
            v_excel_insert_check = False
        else:
            v_excel_insert_check = True
        # 엑셀파일에서 해당 년도 값이 비었으면 항목 추출 실행
        if not insert_chk and v_file_iaif_name == v_table:
            v_cell_row = r[0].row
            print('공시 항목(' + v_file_iaif_name + ') : ' + v_iaif_name)
            main_logic()
            break
    conn.close()
    wb.close()
else:
    print('공시 항목(' + v_table + ') : ' + v_iaif_name)
    main_logic()
    conn.close()
