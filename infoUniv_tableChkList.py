# 개별대학 항목 추출 로직 IAIF_DATA_COMP_US에서 해당 테이블의 DATA가 비교해서 다른 대학만 추출
import time
import sys
import openpyxl as pyxl
import keyboard
import pyperclip
import os.path
import cx_Oracle
import ctypes

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from datetime import datetime
from info_func import excel_insert_check, get_col_title, get_insert_text, cell_extract, reg_d, reg_w,input_func, browser_handle_quit,iaif_path_append

os.environ["NLS_LANG"] = ".AL32UTF8"  # DB 케릭터셋과 일치시킴

# 변수
insert_data, culumn_length, v_cell_row, v_insert_check = [], 0, 0, False
wb, ws = "", ""
col_title_list = []

# v_table: iaifList의 공시파일명
# v_table_org: DATA_COMP_US 테이블에 있는 대학 공시 테이블명, 학과가 존재하는 대학 공시 테이블
v_year = int(input("공시년도 : "))
# v_dtYear = 2018  # 기준년도 : 공시에 따라 info_year 또는 info_year-1

print("생성된 윈도우 창에서 테이블 목록 중 하나를 선택하세요.")
uList = __import__("infoUnivClickList")
time.sleep(0.5)
v_table = uList.check_us_name
# v_table = input("공시 테이블명 : ")  # 파이썬 파일명을 그대로 입력
v_table = v_table.lower()
# v_table_org = input("DATA_COMP 대상 테이블명 : ")  # IAIF_DATA_COMP_US의 테이블명
v_table_org = uList.check_ref_name
v_table_org = v_table_org.upper()

v_read_only = input_func("읽기 전용 여부(Y|N)", "str", ["Y", "N"])  # 읽기 전용 시 병합 체크 로직 처리 못함

v_file_init = input_func("엑셀 파일 초기화 여부(Y/N)", "str", ["Y", "N"])

print(v_table, ":", v_table_org)

if v_table[4:5] == "5":
    univ_grp_type = "A"
elif v_table[4:5] == "7":
    univ_grp_type = "C"
else:
    univ_grp_type = "B"
v_univ_info_except = False

#v_download_folder = r"V:\document\정보공시데이터\PYTHON\excel\down_file"
v_download_folder = r"V:\document\정보공시데이터\PYTHON\excel\edit_file"
v_file_path = v_download_folder + "\\"  # 엑셀 다운로드 경로

time.sleep(0.33)
chrome_option = webdriver.ChromeOptions()
# prefs = {"download.prompt_for_download": True}  # 다운로드 창 열기
prefs = {"download.default_directory": v_download_folder}
chrome_option.add_experimental_option("prefs", prefs)
browser = webdriver.Chrome(options=chrome_option)
browser.maximize_window()
time.sleep(0.66)

browser_handle_quit(browser)

v_col_univ_info = [
    "학교종류"
    , "설립구분"
    , "지역"
    , "상태"
]

# 공시 파일 내 변수 가져오기
iaif_path_append()
v_mod = __import__(v_table)

v_total_col = v_mod.total_col
v_iaif_name = v_mod.iaif_name
print(v_iaif_name)
v_iaif_path = None
try:
    v_iaif_path = __import__(v_table).iaif_path
except AttributeError:
    pass

# 대학원 취업 현황의 전문대학원과 특수대학원은 따로 공시되지만 같은 공시 테이블을 사용하므로
try:
    v_table = v_mod.table_name
except AttributeError:
    pass

try:
    v_dtyear_idx = v_mod.dtyear_idx
except AttributeError:
    v_dtyear = v_year - v_mod.dtyear_num
    v_dtyear_idx = -1  # 공시를 조회한 항목에서 넣어야 할 때, 그 외는 dtyear_num값 만큼 평가년도에서 감소

v_column_start = v_mod.column_start_col

# 공시항목 크롤링 처리 로직
# browser.get("http://academyinfo.go.kr/index.do")

# soup = BeautifulSoup(browser.page_source, 'html.parser')


# 페이지 내 클릭에 의한 다음 루트까지의 요소 검색 지연 처리
def page_flow(browser_element):
    v_check_btn = False
    while not v_check_btn:
        sel_element = browser_element
        if sel_element:
            sel_element[0].click()
            v_check_btn = True
        else:
            v_check_btn = False
        time.sleep(0.4)


def insert_logic(m_mod, m_univ_nm, m_univ_id):  # 테이블 insert 처리
    print(insert_data[:5])
    # show_data(insert_data)
    v_table_cols = ""
    if v_cols_order == "":
        v_table_cols = """(""" + ",".join(v_mod.table_cols) + """)""" if v_mod.table_cols else ""
    elif v_cols_order == "2":
        v_table_cols = """(""" + ",".join(v_mod.table_cols2) + """)""" if v_mod.table_cols2 else ""
    elif v_cols_order == "3":
        v_table_cols = """(""" + ",".join(v_mod.table_cols3) + """)""" if v_mod.table_cols3 else ""

    insert_sql = "INSERT INTO " + v_table + v_table_cols + " VALUES ('" + str(v_year) + "','" + str(v_dtyear) + "','" +\
                 m_univ_nm + "'," + get_insert_text(column_length) + str(m_univ_id) + "," + m_mod.insert_last_col + ")"
    cursor_ins = conn.cursor()
    try:
        # insert_data = unicode(insert_data, "euc-kr").encode("utf-8")
        cursor_ins.executemany(insert_sql, insert_data)
        cursor_ins.close()
        conn.commit()
        print(v_table + " " + m_univ_nm + " insert success!")

        if os.path.isfile(v_new_file_name):
            os.remove(v_new_file_name)
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(insert_sql)
        print(error.code)
        print(error.message)
        print(error.context)
        cursor_ins.close()
        conn.rollback()
        conn.close()
        wb.close()
        sys.exit()


def insert_msg(m_mod, u_nm, u_id):  # 공시와 테이블 비교에 따른 구분
    # 기데이터 삭제
    cursor_del = conn.cursor()
    cursor_del.execute("DELETE FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) + "' AND UNIV_ID = " + str(u_id))
    time.sleep(2)
    cursor_del.close()
    time.sleep(0.33)
    insert_logic(m_mod, u_nm, u_id)


# 대학 공시 검색 --start
def univ_search(chk_file_exist, u_code, u_nm, u_id):
    browser.get("http://academyinfo.go.kr/popup/pubinfo1690/list.do?schlId="+str(u_code))

    page_flow(browser.find_elements_by_xpath("//li[@class='ui-tab']/a[text()='공시정보']"))
    time.sleep(0.5)
    page_flow(browser.find_elements_by_xpath("//li[@class='ui-tab']/a[contains(text(), '전체목록')]"))
    time.sleep(0.7)
    # 공시항목 클릭
    iaif_check = False
    global v_insert_check
    v_insert_check = True

    excel_text = ""
    v_count2 = 0
    if v_iaif_path:
        level_iaif = []
        for idx, val in enumerate(v_iaif_path):
            if idx == 0:
                level_iaif = browser.find_elements_by_xpath("//span[@class='text-list-title'][contains(text(), '" +
                                                            val + "')]")
            elif idx == len(v_iaif_path) - 1:
                level_iaif = level_iaif[0].find_element_by_xpath("../ul//span[contains(text(), '" + val + "')]")
            else:
                level_iaif = level_iaif[0].find_element_by_xpath("../ul//span[contains(text(),'" + val + "')]")
            time.sleep(0.33)
        level_iaif = level_iaif.find_element_by_xpath("./following-sibling::button[contains(text(), '" +
                                                      str(v_year) + "')]")
        level_iaif.click()
    else:
        while not iaif_check:
            # depth-text 클래스 span 태그로 먼저 찾음. 없으면 상위 태그인 text-list-title 클래스 span 태그를 찾는다.
            element = browser.find_elements_by_xpath("//span[@class='depth-text'][contains(text(), '" +
                                                     v_iaif_name + "')]")
            if not element:
                element = browser.find_elements_by_xpath("//span[@class='text-list-title'][contains(text(), '" +
                                                         v_iaif_name + "')]")
            if element:
                try:
                    element = element[0].find_element_by_xpath("./following-sibling::button[contains(text(), '" +
                                                               str(v_year) + "')]")
                    v_insert_check = True
                    element.click()
                except NoSuchElementException:
                    v_insert_check = False
                    excel_text = "현재년도 미공시"
                break
            if v_count2 == 12:
                sys.exit()

            v_count2 += 1
            time.sleep(0.33)

    time.sleep(0.33)

    # 해당년도 항목을 찾았으면 다음 로직을 수행하고, 찾지 못했으면 다음 대학으로 이동
    if v_insert_check:
        browser.switch_to.window(browser.window_handles[1])

        # 공시 항목 데이터 창 open 후, 데이터 불러오기 까지 지연 시키기
        v_check = False
        v_count3 = 0
        while not v_check and v_count3 <= 540:  # 540 : 약 3분
            bs_soup = BeautifulSoup(browser.page_source, 'html.parser')
            check_elmt = bs_soup.find("div", {"id": "UbiHTMLViewer_preview_1"})
            if check_elmt:
                v_check = True

            v_count3 += 1
            time.sleep(0.33)

        # 현재 공시와 저장된 공시의 칼럼 비교 : ext_info_columns.get_col_title
        global col_title_list, cur_title_list
        if len(col_title_list) == 0:
            cur_title_list = get_col_title(browser)

        try:
            global v_cols_order
            if cur_title_list == v_mod.total_col:
                v_cols_order = ""
            elif cur_title_list == v_mod.total_col2:
                v_cols_order = "2"
            elif cur_title_list == v_mod.total_col3:
                v_cols_order = "3"
        except AttributeError:
            ctypes.windll.user32.MessageBoxW(0, str(v_year) + "년도 공시의 제목 열이 기존 제목 열과 다릅니다. 확인 바랍니다.", "", 4096)
            conn.close()
            sys.exit()
        else:
            col_title_list = cur_title_list
        # ###

        if not chk_file_exist:
            datacheck = browser.find_elements_by_xpath(
                "//div[@class='textitem UbiHTMLViewer_previewpage_1font_1 UbiHTMLViewer_previewpage_1color_f_0']")
            if len(datacheck) == 0:
                datacheck = browser.find_elements_by_xpath(
                    "//div[@class='textitem UbiHTMLViewer_previewpage_1font_0_0 UbiHTMLViewer_previewpage_1color_f_0_0']")
            print("데이터 수 : "+str(len(datacheck)))

            global v_save_file_name, v_new_file_name, insert_data, column_length
            v_save_file_name, v_new_file_name = "", ""
            insert_data = []  # table insert data
            column_length = 0  # insert 행의 컬럼 길이

            if len(datacheck) > 1:
                time.sleep(0.3)
                page_flow(browser.find_elements_by_xpath("//td[@id='UbiHTMLViewerUbiToolbar_SaveButton']/input"))  # 저장버튼 클릭
                time.sleep(0.7)
                v_count3 = 0
                v_check = False
                this_year = datetime.today().year
                while not v_check:  # 30초
                    v_save_file_list = os.listdir(v_file_path)
                    if len(v_save_file_list) > 0:
                        for idx, file in enumerate(v_save_file_list):
                            if str(this_year) in file and not file.endswith(".crdownload"):
                                time.sleep(0.5)
                                v_save_file_name = os.listdir(v_file_path)[idx]
                                v_check = True
                                print(v_save_file_name + " saved")
                                v_new_file_name = v_file_path + u_nm + ".xlsx"
                                os.rename(v_file_path + v_save_file_name, v_new_file_name)
                                break

                    v_count3 += 1
                    if v_count3 > 100:
                        print(v_iaif_name + "이 다운로드 되지 않았거나 여러 파일이 존재합니다.")
                        wb.close()
                        browser.quit()
                        sys.exit()
                    time.sleep(0.3)

                # file edit : 바로 파일을 열 경우 cumstomWidth error 발생하므로 전체 복사 저장하는 과정 추가
                time.sleep(1.5)
                if v_check:
                    os.startfile(v_new_file_name)
                    time.sleep(4)
                    v_check = False
                    v_count = 0
                    while not v_check:
                        keyboard.press_and_release("control+c")
                        time.sleep(0.33)
                        cell_text = pyperclip.paste()
                        if str(v_year) in cell_text or "info" in cell_text:
                            time.sleep(0.8)
                            keyboard.press_and_release("control+s")
                            time.sleep(0.5)
                            keyboard.press_and_release("alt+f+x")
                            time.sleep(0.8)
                            v_check = True
                            print("file edited")

                        v_count += 1
                        if v_count > 33:
                            print("저장된 파일을 실행시키는데 실패하였습니다.")
                            conn.close()
                            wb.close()
                            browser.quit()
                            sys.exit()
                        time.sleep(0.33)

                else:
                    print("공시 데이터가 없습니다.")
                    ws.cell(v_cell_row, 3).value = "공시 데이터 없음"
                    return

                time.sleep(0.33)

        else:
            v_new_file_name = v_file_path + u_nm + ".xlsx"

    # 다운받은 파일을 열어서 데이터 추출
    global wb_info, ws_info
    v_read_only_chk = True if v_read_only == "Y" else False
    wb_info = pyxl.load_workbook(filename=v_new_file_name, read_only=v_read_only_chk)
    ws_info = wb_info.active  # or wb.active : 활성화된 시트

    v_compare_row_num = v_mod.excel_data_start_row
    merged_cell_list = []  # 데이터 중 병합이 있는 범위를 담는 list

    # 데이터 중 칼럼의 병합은 불가. 로우의 병합만 존재.
    if v_read_only == "N":
        for m_row in ws_info.merged_cells.ranges:
            v_cell_range = cell_extract(m_row)
            if len(v_cell_range[0]) > 1:
                print("병합된 열이 Z보다 큽니다. 확인 필요")
                conn.close()
                wb.close()
                browser.quit()
                sys.exit()

            elif v_cell_range[1] >= v_compare_row_num + 1:
                # 행 : 셀은 1부터 시작하므로 +1
                # 열 : 시작 출발 지정 알파벳 열부터 포함
                merged_cell_list.append(v_cell_range)  # [칼럼 알파벳, 병합시작row, 병합종료row]

    ws_row_list = list(ws_info.rows)
    v_last_row_except_num = 1 if ws_row_list[len(ws_row_list) - 1][0].value is None else 0
    print('엑셀 데이터 수 :', len(ws_row_list))
    for idx, row in enumerate(ws_row_list[v_compare_row_num:len(ws_row_list) - v_last_row_except_num]):
        global item_list
        item_list = []
        # 행 제외 데이터에 포함되면 해당 행 부분은 insert하지 않고 다음 반복으로 이동
        # us 파일에 delete_row_compare_col_index 변수 선언
        try:
            if (str(row[v_mod.delete_row_compare_col_index[0]].value) in v_mod.delete_row_compare_col_index[1]
               or row[v_mod.delete_row_compare_col_index[0]].value is None):
                continue
        except AttributeError:
            pass
        except IndexError:  # 행의 데이터 수가 부족한 행은 일반적인 데이터 행이 아니므로 제외
            continue

        # 불필요한 열 병합으로 인한 빈 칼럼에 해당하는 데이터 제외하기
        # us 파일에 delete_col_index_list 변수 선언
        try:
            row = [el for no, el in enumerate(row) if no not in v_mod.delete_col_index_list]
        except AttributeError:
            pass

        for d_idx, d_value in enumerate(list(row)[v_column_start:]):  # 지정된 시작 열부터 포함
            global row_value
            row_value = str(d_value.value)

            if merged_cell_list:
                if row_value is None or row_value == "" or row_value == "None":
                    addr_w = reg_w.search(d_value.coordinate).group()
                    addr_d = int(reg_d.search(d_value.coordinate).group())
                    for m_arr in merged_cell_list:
                        if addr_w == m_arr[0] and m_arr[1] <= addr_d <= m_arr[2]:  # 병합된 셀
                            row_value = insert_data[len(insert_data) - 1][d_idx]  # 이전 행의 해당 인덱스의 요소 할당
                            # print('이전 값 : ', row_value)
                            break
            else:
                if d_idx in v_mod.merge_check_cols:
                    # 데이터 행에 대한 시작 열을 지정해서 가져오기 때문에 그 수치만큼 더하여야 한다.(univ_nm_idx)
                    if row_value in ["None", ""]:
                        row_value = insert_data[len(insert_data) - 1][d_idx]  # 이전 행의 해당 인덱스의 요소 할당

            row_value = row_value.replace(",", "")

            try:
                if d_idx == v_mod.data_change[0]:
                    row_value = v_mod.data_change[2] if row_value == v_mod.data_change[1] else row_value
            except AttributeError:
                pass

            item_list.append(row_value)

        if len(item_list) > 0:
            # insert_year = [str(v_year), str(v_dtyear)]
            insert_data.append(item_list)
        else:
            continue

        if len(insert_data) == 1:
            # print(insert_data[0])
            column_length = len(insert_data[0])

    wb_info.close()
    insert_msg(v_mod, u_nm, u_id)
    browser.close()
    browser.switch_to.window(browser.window_handles[0])
    excel_insert_check(ws, v_cell_row, v_insert_check, excel_text)
    wb.save(file_path + v_iaif_name + v_file_add_text + ".xlsx")

    time.sleep(0.33)
# 대학 공시 검색 --end


# oracle 연동
dsn = cx_Oracle.makedsn("61.81.234.137", 1521, "COGDW")
conn = cx_Oracle.connect("dusd", "dusd$#@!", dsn)
cursor = conn.cursor()
cursor.execute("SELECT DISTINCT B.UNIV_NM, B.UNIV_ID" +
               "  FROM IAIF_DATA_COMP_US A, KUES0004 B" +
               " WHERE A.UNIV_ID = B.UNIV_ID" +
               "   AND TABLE_NM = '" + v_table_org + "'" +
               "   AND INFO_YEAR = '" + str(v_year) + "'" +
               "   AND DATA_CHK = 'NOT EQUAL'" +
               " ORDER BY UNIV_NM")
result = cursor.fetchall()
cursor.close()

# result는 list 안에 각 요소가 하나의 요소를 가진 tuple로 되어 있기 때문에 tuple을 요소로 전환한다.
# for idx, val in enumerate(result):
#     result[idx] = val

print(result)

v_file_add_text = "_전문대학" if univ_grp_type == "C" else ""
file_path = "V:\document\정보공시데이터\PYTHON\excel\\"
wb = pyxl.load_workbook(file_path+v_iaif_name+v_file_add_text+".xlsx")
ws = wb['Sheet1']  # or .active

if v_file_init == "Y":
    for r in list(ws.rows)[1:]:
        ws.cell(row=r[0].row, column=3).value = ""
        ws.cell(row=r[0].row, column=4).value = ""
    wb.save(file_path+v_iaif_name+v_file_add_text+".xlsx")
    print(v_iaif_name+" 엑셀 파일 초기화")

v_insert_univ_list = ""

for r_no, r_val in enumerate(result):  # 개별대학 추가 확보 필요 대학별로 공시 개별대학 엑셀 파일에서 해당 대학 코드 찾기
    univ_nm = r_val[0]
    univ_id = r_val[1]
    for r in list(ws.rows)[1:]:
        if univ_nm == r[0].value:
            v_insert_univ_list += str(univ_id) + ","
            if r[3].value != "O":
                print(univ_nm+" : start")
                v_cell_row = r[0].row
                ws.cell(row=v_cell_row, column=3).value = "1"
                if os.path.isfile(v_file_path + univ_nm + ".xlsx"):
                    univ_search(True, r[1].value, univ_nm, univ_id)
                else:
                    univ_search(False, r[1].value, univ_nm, univ_id)
                print('-------------------')
                break


print("UNIV_ID :", v_insert_univ_list)
wb.save(file_path+v_iaif_name+v_file_add_text+".xlsx")
wb.close()
browser.quit()
sys.exit()
