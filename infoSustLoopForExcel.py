# 학과별 정보공시 업로드 처리 : 정보공시 재업데이트 관리_PYTHON 학과 시트 LOOP
# iaif_file에서 table_col에 part_id 위치 주의
import ctypes
import time
import sys
import os
import keyboard
import openpyxl as pyxl
import pyperclip
import cx_Oracle
import shutil

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException
from datetime import datetime

from info_func import get_col_title, get_insert_text, element_check, is_univ_nm_check, cell_extract, reg_d, reg_w, \
    file_resave, input_func, browser_handle_quit, fileDLcheck, iaif_path_append
import infoClickYearly

os.environ["NLS_LANG"] = ".AL32UTF8"  # DB 케릭터셋과 일치시킴

wb, ws, wb_info, ws_info, v_cell_row, browser, bs_soup = "", "", "", "", 0, "", ""
dsn, conn = "", ""
result, insert_data, column_length = "", "", ""
v_year, v_table, univ_grp_type, v_mod = "", "", "", ""
v_cols_order = ""  # 항목 칼럼이 상이할 시 col, col2, ...에서 지정하는 순서
v_dtyear, v_sheet_name = "", ""

print("오류 발생 시 down_file 내 파일이 존재하면 삭제하세요.")
print(infoClickYearly.v_down_year)

part_list = ["1", "2", "3", "4", "5"]
part_name = {"1": "인문ㆍ사회계열", "2": "자연과학계열", "3": "공학계열", "4": "예ㆍ체능계열", "5": "의학계열"}

v_univgrp1 = input_func("대학종류(A|B|C) : ", "str", ["A", "B", "C"])  # 대학|전문대학
if v_univgrp1.upper() == "A":
    v_univgrp1 = "대학"
elif v_univgrp1.upper() == "C":
    v_univgrp1 = "전문대학"
else:
    v_univgrp1 = "대학원"

v_sheet_name = "학과별_" + v_univgrp1

v_download_folder = r"V:\document\정보공시데이터\PYTHON\excel\down_file"
v_file_path = v_download_folder + "\\"  # 엑셀 다운로드 경로
v_file_path_down_only = v_download_folder + "2\\"  # 다운로드 전용 학과파일 엑셀 저장 폴더 경로


# 페이지 내 클릭에 의한 다음 루트까지의 요소 검색 지연 처리
def page_flow(browser_element):
    v_count2 = 0
    v_check_btn = False
    while not v_check_btn:
        sel_element = browser_element
        if sel_element:
            sel_element[0].click() if isinstance(sel_element, list) else sel_element.click()
            v_check_btn = True
        else:
            v_check_btn = False

        if v_count2 == 13:
            print("요소 찾기 시간 초과")
            sys.exit()

        v_count2 += 1
        time.sleep(0.33)


def insert_logic(m_mod, m_part_id, ):  # 테이블 insert 처리
    print(insert_data[:10])
    # show_data(insert_data)
    v_table_cols = ""
    if v_cols_order == "":
        v_table_cols = "(" + ",".join(v_mod.table_cols) + ")" if v_mod.table_cols else ""
    elif v_cols_order == "2":
        v_table_cols = "(" + ",".join(v_mod.table_cols2) + ")" if v_mod.table_cols2 else ""
    elif v_cols_order == "3":
        v_table_cols = "(" + ",".join(v_mod.table_cols3) + ")" if v_mod.table_cols3 else ""

    insert_sql = "INSERT INTO " + v_table + v_table_cols + " VALUES ('" + str(v_year) + "'," + \
                 get_insert_text(column_length) + m_mod.insert_last_col + ")"
    cursor_ins = conn.cursor()
    try:
        # insert_data = unicode(insert_data, "euc-kr").encode("utf-8")
        cursor_ins.executemany(insert_sql, insert_data)
        cursor_ins.close()
        conn.commit()
        print(v_table + " " + part_name[m_part_id] + " insert success!")
        ws.cell(v_cell_row, v_year_index_in_file + 1).value = "O"  # 엑셀 칼럼은 1부터 시작하므로 년도 추출 변수에 1을 더함
        wb.save("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")

        if os.path.isfile(v_file_path + v_new_file_name):
            os.remove(v_file_path + v_new_file_name)
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(insert_sql)
        print(error.code)
        print(error.message)
        print(error.context)
        cursor_ins.close()
        conn.rollback()
        conn.close()
        wb.close()
        sys.exit()


def insert_msg(value, m_mod, m_part_id):  # 공시와 테이블 비교에 따른 구분
    if value == 1:  # 예 클릭
        try:
            if v_mod.delete_except:
                print("해당 항목은 Delete 제외 공시입니다.")
                pass
        except AttributeError:
            # 기데이터 삭제
            cursor_del = conn.cursor()
            cursor_del.execute(
                "DELETE FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) + "' AND PART_ID = " +
                m_part_id)
            time.sleep(2)
            cursor_del.close()
        time.sleep(0.33)
        insert_logic(m_mod, m_part_id)
    else:
        browser.quit()
        conn.rollback()
        conn.close()
        sys.exit()


def main_logic(m_iaif_name, m_part_id):
    print("-- 시작 --")
    # 파라미터
    global v_year
    v_year = infoClickYearly.v_down_year
    global v_table
    v_table = m_iaif_name
    v_table = v_table.lower()
    global univ_grp_type
    univ_grp_type = ""  # 태그의 option value - 01[전문대학], 02[대학], 03[대학원]

    if v_table[4:5] == "5":
        univ_grp_type = "02"
    elif v_table[4:5] == "7":
        univ_grp_type = "01"
    else:
        univ_grp_type = "03"

    part_id = str(m_part_id)

    # 공시 파일 내 변수 가져오기
    iaif_path_append()
    global v_mod
    v_mod = __import__(v_table)

    v_iaif_name = v_mod.iaif_name

    v_iaif_ord_name = None
    try:
        v_iaif_ord_name = __import__(v_table).iaif_ord_name
    except AttributeError:
        pass

    # 대학원 취업 현황의 전문대학원과 특수대학원은 따로 공시되지만 같은 공시 테이블을 사용하므로
    # 뒤에 오는 특수대학원에 공시 테이블명을 두어 다른 파일명이지만 같은 테이블에 insert 되도록 처리
    try:
        v_table = v_mod.table_name
    except AttributeError:
        pass

    global v_dtyear
    try:
        v_dtyear_idx = v_mod.dtyear_idx
    except AttributeError:
        v_dtyear = int(v_year) - int(v_mod.dtyear_num)
        v_dtyear_idx = -1  # 공시를 조회한 항목에서 넣어야 할 때, 그 외는 dtyear_num값 만큼 평가년도에서 감소

    if v_iaif_ord_name:
        # element[0].send_keys(v_iaif_org_name)
        browser.get('http://academyinfo.go.kr/search/search.do?kwd="' + v_iaif_ord_name + '"&schlKnd=' + univ_grp_type)
    else:
        # element[0].send_keys(v_iaif_name)
        browser.get('http://academyinfo.go.kr/search/search.do?kwd="' + v_iaif_name + '"&schlKnd=' + univ_grp_type)

    # element[0].send_keys(Keys.ENTER)
    time.sleep(1.5)

    # 공시 데이터 검색
    global element
    try:
        element = element_check(browser.find_element_by_xpath(
            "//table[@class='tbl-col']/tbody[@id='targetDiv']//td[contains(text(),'" +
            v_iaif_name + "')]"))
    except:
        try:
            element = element_check(browser.find_element_by_xpath(
                "//table[@class='tbl-col']/tbody[@id='targetDiv']//td[contains(text(),'" +
                v_iaif_ord_name + "')]"))
        except:
            print("공시 항목을 찾을 수 없습니다.")
            sys.exit()

    # 공시명 우측 '학교별평균값' select box 클릭
    actions = ActionChains(browser)
    # xpath상에서 index는 1부터 시작
    actions.move_to_element(element.find_element_by_xpath("..//span[contains(@class, 'ui-select-wrap')][1]"))
    actions.click()
    actions.perform()
    time.sleep(0.33)

    # 학과별 클릭
    page_flow(browser.find_element_by_xpath("//div[contains(@class, 'ui-selectmenu-open')]//div[text()='학과별']"))

    # 계열 select box 선택하기
    page_flow(element.find_element_by_xpath("..//span[@id='pgmStags']"))

    # 계열 선택하기
    page_flow(browser.find_element_by_xpath(
        "//div[contains(@class, 'ui-selectmenu-open')]//div[text()='" + part_name[part_id] + "']"))

    # 년도 클릭
    try:
        page_flow(element.find_element_by_xpath("..//button[@data-svy_yr='" + str(v_year) + "']"))
    except NoSuchElementException:
        print(str(v_year), "년도 데이터가 없습니다.")
        ws.cell(v_cell_row, v_year_index_in_file + 1).value = "미공시"
        return 0

    browser.switch_to.window(browser.window_handles[1])  # 생성된 탭 페이지 이동
    global bs_soup
    bs_soup = BeautifulSoup(browser.page_source, 'html.parser')

    # 공시 항목 데이터 창 open 후, 데이터 불러오기 까지 지연 시키기
    v_check = False
    v_count3 = 0
    while not v_check and v_count3 <= 540:  # 540 : 약 3분
        bs_soup = BeautifulSoup(browser.page_source, 'html.parser')
        check_elmt = bs_soup.find("div", {"id": "UbiHTMLViewer_preview_1"})
        if check_elmt:
            v_check = True

        v_count3 += 1
        time.sleep(0.33)

    # 현재 공시와 저장된 공시의 칼럼 비교 : ext_info_columns.get_col_title
    cur_title_list = get_col_title(browser)

    try:
        global v_cols_order
        if cur_title_list == v_mod.total_col:
            v_cols_order = ""
        elif cur_title_list == v_mod.total_col2:
            v_cols_order = "2"
        elif cur_title_list == v_mod.total_col3:
            v_cols_order = "3"
    except AttributeError:
        ctypes.windll.user32.MessageBoxW(0, str(v_year) + "년도 공시의 제목 열이 기존 제목 열과 다릅니다. 확인 바랍니다.", "", 4096)
        conn.close()
        sys.exit()
    # ###

    # 모든 페이지를 불러왔는지를 확인
    v_check = False
    v_count3 = 0
    while not v_check:
        page_check = browser.find_elements_by_xpath("//div[@id='UbiHTMLViewerUbiToolbarButton_TotalPageText']")[0].text
        if "+" not in page_check:
            v_check = True
        elif v_count3 >= 540:
            ctypes.windll.user32.MessageBoxW(0, str(v_year) + "모든 페이지를 불러오지 못했습니다. 확인 바랍니다.", "", 4096)
            conn.close()
            sys.exit()

        v_count3 += 1
        time.sleep(0.33)

    # 열에서 학교 체크
    bs_soup2 = BeautifulSoup(browser.page_source, 'html.parser')

    global insert_data
    insert_data = []  # table insert data
    global column_length
    column_length = 0  # insert 행의 컬럼 길이

    my_element4 = bs_soup2.find("div", {"id": "UbiHTMLViewer_previewpage_1"})
    my_element4 = my_element4.find_all("div", {"class": "UbiHTMLViewer_previewpage_1color_b_2"})

    v_univ_title_index = None
    for idx, val in enumerate(my_element4):
        if val.get_text() in ["학교명", "학교", "대학명", '대학원명']:
            v_univ_title_index = idx
            break

    datacheck = browser.find_elements_by_xpath(
        "//div[@class='textitem UbiHTMLViewer_previewpage_1font_1 UbiHTMLViewer_previewpage_1color_f_0']")
    if len(datacheck) == 0:
        datacheck = browser.find_elements_by_xpath(
            "//div[@class='textitem UbiHTMLViewer_previewpage_1font_0_0 UbiHTMLViewer_previewpage_1color_f_0_0']")

    global v_save_file_name, v_new_file_name
    v_save_file_name, v_new_file_name = "", ""
    print("데이터 수 : " + str(len(datacheck)))
    if len(datacheck) > 1:
        time.sleep(0.3)
        v_save_file_list = os.listdir(v_file_path)
        if len(v_save_file_list) >= 1 and v_save_file_list[0] != "Thumbs.db":
            print(v_save_file_list)
            print("PYTHON>excel>down_file 폴더에 파일이 존재합니다.")
            conn.close()
            wb.close()
            browser.quit()
            sys.exit()
        else:
            page_flow(browser.find_elements_by_xpath("//td[@id='UbiHTMLViewerUbiToolbar_SaveButton']/input"))  # 저장버튼 클릭

        # file save
        time.sleep(1)
        v_count3 = 0
        v_check = False
        this_year = datetime.today().year
        while not v_check:  # 100초
            v_save_file_list = os.listdir(v_file_path)
            if len(v_save_file_list) > 0:
                for idx, file in enumerate(v_save_file_list):
                    if str(this_year) in file and not file.endswith(".crdownload"):
                        time.sleep(2.2)
                        v_save_file_name = os.listdir(v_file_path)[idx]
                        if not fileDLcheck(v_file_path + v_save_file_name):
                            conn.close()
                            wb.close()
                            browser.quit()
                            sys.exit()

                        v_check = True
                        print(v_save_file_name + " saved")
                        v_new_file_name = v_iaif_name + "_" + str(v_year) + "_" + part_name[str(v_file_part_id)] + ".xlsx"
                        time.sleep(0.3)
                        os.rename(v_file_path + v_save_file_name, v_file_path + v_new_file_name)
                        try:
                            if v_mod.download_only:
                                if not file_resave(v_file_path, v_new_file_name, v_year):
                                    conn.close()
                                    wb.close()
                                    browser.quit()
                                    sys.exit()
                                time.sleep(0.6)
                                shutil.move(v_file_path + v_new_file_name, v_file_path_down_only + v_new_file_name)
                                ws.cell(v_cell_row, v_year_index_in_file + 1).value = "다운로드 완료"
                                wb.save("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")
                                return
                        except AttributeError:
                            pass

            v_count3 += 1
            if v_count3 > (400 if v_table == "iaif5406s" else 200):
                print(v_iaif_name + "의 " + part_name[part_id] + "이 다운로드 되지 않았거나 여러 파일이 존재합니다.")
                conn.close()
                wb.close()
                browser.quit()
                sys.exit()
            time.sleep(0.3)

        # file edit : 바로 파일을 열 경우 cumstomWidth error 발생하므로 전체 복사 저장하는 과정 추가
        time.sleep(1.2)
        if v_check:
            if not file_resave(v_file_path, v_new_file_name, v_year):
                conn.close()
                wb.close()
                browser.quit()
                sys.exit()

    else:
        print("공시 데이터가 없습니다.")
        ws.cell(v_cell_row, v_year_index_in_file + 1).value = "데이터 없음"
        wb.save("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")
        return

    time.sleep(0.5)

    # 대학명이 존재하는 칼럼 index 할당 : 일부 공시는 학교종류가 병합될 수 도 있음
    # 보통 데이터는 대학명부터 잘라서 집어넣는데, 대학원(신입생 충원 현황)의 경우 대학원종류 칼럼부터 insert하므로 이를 처리
    try:
        univ_nm_idx = v_mod.main_data_start_index
    except AttributeError:
        univ_nm_idx = v_univ_title_index

    print("file open")
    # 다운받은 파일을 열어서 데이터 추출
    global wb_info, ws_info
    wb_info = pyxl.load_workbook(v_file_path + v_new_file_name)
    ws_info = wb_info.active  # or wb.active : 활성화된 시트

    v_compare_row_num = v_mod.excel_data_start_row
    merged_cell_list = []  # 데이터 중 병합이 있는 범위를 담는 list
    # 데이터 중 칼럼의 병합은 불가. 로우의 병합만 존재.
    for m_row in ws_info.merged_cells.ranges:
        v_cell_range = cell_extract(m_row)
        if v_cell_range[1] >= v_compare_row_num + 1:  # 셀은 1부터 시작하므로 +1
            merged_cell_list.append(v_cell_range)  # [칼럼 알파벳, 병합시작row, 병합종료row]

    v_data_chk_list = []  # 중복 체크를 위한 키 묶음 텍스트(v_data_key_txt) 저장 list
    v_data_dup_list = {}  # 중복되는 데이터 담는 dictionary
    for idx, row in enumerate(list(ws_info.rows)[v_compare_row_num:]):
        item_list = []
        v_data_key_txt = ""  # 키 묶음 텍스트 변수

        # 기준년도 칼럼 추출. (재학생 충원 현황: iaif5108s_2013)
        if v_table in ("iaif5108s_2013", "iaif7108s_2013"):
            v_dtyear = row[v_dtyear_idx].value if v_dtyear_idx > -1 else str(v_dtyear)
        else:
            pass

        for d_idx, d_value in enumerate(list(row)[univ_nm_idx:]):
            global row_value
            row_value = str(d_value.value)

            if merged_cell_list:
                if not row_value:
                    addr_w = reg_w.search(d_value.coordinate).group()
                    addr_d = int(reg_d.search(d_value.coordinate).group())
                    for m_arr in merged_cell_list:
                        if addr_w == m_arr[0] and m_arr[1] <= addr_d <= m_arr[2]:  # 병합된 셀
                            row_value = insert_data[len(insert_data) - 1][d_idx]  # 이전 행의 해당 인덱스의 요소 할당
                            print('이전 값 : ', row_value)
                            break

            if is_univ_nm_check(row_value):
                row_value = row_value.replace(" _", "_")
            elif v_year == 2016 and v_table in ("iaif5200_13", "iaif6200"):
                # 졸업생 취업률의 취업률 칼럼은 16년 이전에는 '-'로 표시되었기 때문에 ''로 수정
                row_value = "" if row_value == "-" else row_value.replace(",", "")
            else:
                # 수치 데이터의 ','(comma) 표시 제거
                try:
                    if float(row_value.replace(",", "")):
                        row_value = row_value.replace(",", "")  # .replace("∙", "ㆍ")
                except ValueError:
                    pass

            # if univ_grp_type == "03":
            #     row_value = is_univ_nm_space_check(row_value)

            # v_mod.data_key_index는 IAIF 파일 변수인데 이는 엑셀 다운로드 데이터 기준 인덱스이기 때문에
            # 현재 루프의 d_idx에서 시작하는 열로 지정한 univ_nm_idx만큼 더해야 한다.
            if (d_idx + univ_nm_idx) in v_mod.data_key_index:
                # row_value가 NoneType일 경우 오류 발생 : 이전 로직에서 None을 반환하였기 때문
                v_data_key_txt += row_value if v_data_key_txt == "" else "/" + row_value

            item_list.append(row_value)

        if v_data_key_txt in v_data_chk_list:
            if v_data_key_txt in v_data_dup_list.keys():
                v_data_dup_list[v_data_key_txt] += 1
            else:
                v_data_dup_list[v_data_key_txt] = 2
        else:
            v_data_chk_list.append(v_data_key_txt)

        # 일부 공시는 테이블이 분리되어 있기 때문에 나눠서 추출한다. (국내외 연구실적, 저역서 실적)
        v_div_list_module, v_data_div_list = [], []
        try:
            if v_mod.data_div_rng:
                # v_cols_order의 순서에 따라 data_div_rng의 번호를 할당
                if v_cols_order == "":
                    v_div_list_module = v_mod.data_div_rng
                elif v_cols_order == "2":
                    v_div_list_module = v_mod.data_div_rng2
                elif v_cols_order == "3":
                    v_div_list_module = v_mod.data_div_rng3

            for idx, value in enumerate(v_div_list_module):
                # print('칼럼 구분 숫자 : '+str(value))
                # print(item_list)
                # print(item_list[univ_nm_idx:value] if idx == 0 else item_list[value[0]:value[1]])
                if idx == 0:
                    v_data_div_list += item_list[:value-univ_nm_idx]
                else:
                    v_data_div_list += item_list[value[0]-univ_nm_idx:value[1]-univ_nm_idx]
        except AttributeError:
            v_data_div_list = None
            pass

        wb_info.close()

        if len(item_list) > 0:
            v_dtyear_element = [str(v_dtyear)]
            insert_data.append(v_dtyear_element + (item_list if not v_data_div_list else v_data_div_list)
                               + list(str(part_id)))
        else:
            continue

        if len(insert_data) == 1:
            # print(insert_data[0])
            column_length = len(insert_data[0])

    # v_mod(공시 항목 파일)에 data_dup_except 변수가 True이면 중복 Pass
    # 해당 변수가 없고, 중복 리스트에 값이 존재하면 중단.
    try:
        if v_mod.data_dup_except:
            pass
    except AttributeError:
        if v_data_dup_list:
            print("중복되는 데이터가 존재합니다.")
            print(v_data_dup_list)
            conn.close()
            wb.close()
            browser.quit()
            sys.exit()
        else:
            pass

    # for elmt in insert_data[:10]:
    #     print(elmt)
    insert_msg(1, v_mod, part_id)
# end-function


if infoClickYearly.v_total_insert_yn == "Y":

    chrome_option = webdriver.ChromeOptions()
    prefs = {"download.default_directory": v_download_folder}
    chrome_option.add_experimental_option("prefs", prefs)
    browser = webdriver.Chrome(options=chrome_option)
    browser.maximize_window()
    time.sleep(1)

    browser_handle_quit(browser)

    browser.get("http://academyinfo.go.kr/index.do")
    time.sleep(1.5)
    browser_handle_quit(browser)

    bs_soup = BeautifulSoup(browser.page_source, 'html.parser')

    # oracle 연동
    dsn = cx_Oracle.makedsn("61.81.234.137", 1521, "COGDW")
    conn = cx_Oracle.connect("dusd", "dusd$#@!", dsn)

    # 정보공시 재업데이트 공시 목록 불러오기
    wb = pyxl.load_workbook("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")
    ws = wb[v_sheet_name]  # or wb.active : 활성화된 시트
    v_year_index_in_file = 2

    # infoClickYearly 파일의 지정 년도가 재업데이트 파일에서의 칼럼의 년도에서의 위치 추출
    for idx, col in enumerate(list(ws.rows)[0]):
        if str(col.value) == str(infoClickYearly.v_down_year):
            v_year_index_in_file = idx
            break

    for r in list(ws.rows)[1:]:
        insert_chk = r[v_year_index_in_file].value
        # 엑셀파일에서 해당 년도 값이 비었으면 항목 추출 실행
        if not insert_chk:
            v_cell_row = r[0].row
            global v_file_iaif_name
            print(r[0].value)
            print(r[1].value)
            v_file_iaif_name = r[0].value
            v_file_part_id = r[1].value
            print("공시 항목 : " + v_file_iaif_name + ", 계열 : " + part_name[str(v_file_part_id)])

            if main_logic(v_file_iaif_name, v_file_part_id) != 0:
                browser.close()
                browser.switch_to.window(browser.window_handles[0])
                browser.delete_all_cookies()


conn.close()
wb.close()
browser.quit()
