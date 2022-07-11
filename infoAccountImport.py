# 재정알리미 다운로드 후 엑셀자료를 추출하여 import하는 파일
import time
import sys
import os
import openpyxl as pyxl
import xlrd
import cx_Oracle
import tkinter as tk
import tkinter.ttk as ttk
from info_func import get_insert_text, iaif_path_append

os.environ["NLS_LANG"] = ".AL32UTF8"  # DB 케릭터셋과 일치시킴

v_start_row = 2
v_univ_cols = [1, 2, 3, 4, 6]  # F열은 병합처리되어 제외


def box_close(event=None):
    global v_file_name
    v_file_name = comboSel.get()
    combo.destroy()


def go_exit():
    sys.exit()


v_year = input("폴더 년도 : ")

# 본인 컴퓨터의 드라이브 명으로 지정(V: -> ?)
v_file_path = "V:\\document\\정보공시데이터\\대학\\사립대학회계정보시스템 추출파일\\자금예산서_교비회계(예산)\\"

combo = tk.Tk()
combo.geometry("450x120")
combo.title("직종 분류 선택")
label1 = tk.Label(combo, text="파일명: ", width=12).grid(row=1, column=0)
comboSel = ttk.Combobox(combo, values=os.listdir(v_file_path + str(v_year)), width=35)
comboSel.grid(row=1, column=1)
comboSel.current(0)
combo.bind("<Return>", box_close)
btn = tk.Button(combo, text="추출 실행", command=box_close).grid(row=3, column=1)
btn2 = tk.Button(combo, text="종료", command=go_exit).grid(row=4, column=1)
combo.lift()
combo.focus_force()
combo.mainloop()

# input("파일명(확장자포함) : ")
print(v_file_name)
v_table = input("테이블명 : ")
v_table = v_table.lower()

v_finan_num = input("1.수입, 2.지출(숫자 입력) : ")
v_finan_gbn = {"1": "수입", "2": "지출"}
v_finan_name = v_finan_gbn[v_finan_num]

v_sheet_name = input("시작할 시트명(처음 Enter) : ")

# 공시 파일 내 변수 가져오기
iaif_path_append()
v_mod = __import__(v_table)

v_sheet_double_chk = v_mod.sheet_double_chk
v_dt_year_num = v_mod.dt_year_num
v_finan_item_list = v_mod.finan_gbn[v_finan_name]

v_file_full_path = v_file_path + str(v_year) + "\\" + v_file_name

# oracle 연동
dsn = cx_Oracle.makedsn("61.81.234.137", 1521, "COGDW")
conn = cx_Oracle.connect("dusd", "dusd$#@!", dsn)

v_table_cols = "(" + ",".join(v_mod.table_cols) + ")" if v_mod.table_cols else ""

insert_sql = "INSERT INTO " + v_table + v_table_cols + " VALUES (" + get_insert_text(len(v_mod.table_cols)-2) + v_mod.insert_last_col + ")"
cursor = conn.cursor()

# 정보공시 재업데이트 공시 목록 불러오기
wb = xlrd.open_workbook(v_file_full_path)

v_univ_name = []
insert_data = []
v_univ_data = {
    "univ_data1": [],
    "univ_data2": [],
    "univ_data3": [],
    "univ_data4": [],
    "univ_data5": []
}

v_sheet_start_chk = False

for sheet_idx, sheet_name in enumerate(wb.sheet_names()):
    if v_sheet_name == "" or v_sheet_start_chk:
        pass
    else:
        if sheet_name == v_sheet_name:
            v_sheet_start_chk = True
        else:
            continue

    print("<<", sheet_name, "start >>")
    ws = wb.sheet_by_name(sheet_name)  # or wb.active : 활성화된 시트

    if v_sheet_double_chk and sheet_idx % 2 == 0:  # 짝수 번(홀수 페이지) 시작 시 reset
        pass
    else:
        v_univ_name = []
        insert_data = []
        v_univ_data = {
            "univ_data1": [],
            "univ_data2": [],
            "univ_data3": [],
            "univ_data4": [],
            "univ_data5": []
        }

    v_start_row = 0 if v_sheet_double_chk and sheet_idx % 2 == 1 else v_start_row

    for r_idx in range(v_start_row, ws.nrows):
        if r_idx == v_start_row:
            if v_sheet_double_chk and sheet_idx % 2 == 1:
                pass
            else:
                for c_no in v_univ_cols:
                    try:
                        if len(ws.cell(r_idx, c_no).value) > 3:
                            v_univ_name.append(ws.cell(r_idx, c_no).value + "_본교")
                        else:
                            continue
                    except:
                        break
        else:
            v_item = ws.cell(r_idx, 0).value
            v_mod_item = v_item.lstrip()
            v_size_item = len(v_item)
            v_size_mod_item = len(v_mod_item)
            v_level = int((v_size_item - v_size_mod_item)/2)  # 2개 공백이 한개 레벨임
            v_level_item_name = str(v_level) + ". " + v_mod_item

            for c_idx, c_univ in enumerate(v_univ_name):
                v_univ_data["univ_data" + str(c_idx+1)].append([
                    v_year, str(int(v_year) - v_dt_year_num), v_finan_name, c_univ,
                    v_mod_item, v_level, v_level_item_name, v_finan_item_list[v_level_item_name][2],
                    None if ws.cell(r_idx, v_univ_cols[c_idx]).value == "null" else ws.cell(r_idx, v_univ_cols[c_idx]).value
                ])

    if v_sheet_double_chk and sheet_idx % 2 == 0:  # 짝수 번(홀수 페이지) 시작 시 reset
        continue
    else:
        try:
            for key in v_univ_data.keys():
                insert_data = insert_data + v_univ_data[key]
            print(v_univ_name)

            cursor.executemany(insert_sql, insert_data)
            conn.commit()
            print(sheet_name + " insert success!")
        except cx_Oracle.DatabaseError as e:
            error, = e.args
            print(insert_sql)
            print(error.code)
            print(error.message)
            print(error.context)
            print(sheet_name + " error!")
            cursor.close()
            conn.rollback()
            conn.close()
            wb.release_resources()
            sys.exit()


conn.close()
wb.release_resources()
