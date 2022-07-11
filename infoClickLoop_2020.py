# 대학별 정보공시 다운로드
import ctypes
import time
import sys
import re
import os
import openpyxl as pyxl
import cx_Oracle

from bs4 import BeautifulSoup
from selenium import webdriver
from datetime import datetime
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementNotInteractableException

from info_func import get_col_title, get_insert_text, input_func, browser_handle_quit, iaif_path_append
import infoClickYearly

os.environ["NLS_LANG"] = ".AL32UTF8"  # DB 케릭터셋과 일치시킴

print('infoClickYearly에서 공시년도를 맞게 입력했는지 확인하세요.')
print('년도 : ', infoClickYearly.v_down_year)

v_univgrp1 = input_func("대학종류(A|C)", "str", ["A", "C"])  # 대학|전문대학
v_univgrp1 = "대학" if v_univgrp1.upper() == "A" else "전문대학"
v_sheet_name = v_univgrp1

wb, ws, v_cell_row, browser = "", "", 0, ""
dsn, conn = "", ""
result, insert_data, column_length = "", "", ""
v_year, v_table, univ_grp_type, v_mod = "", "", "", ""
v_cols_order = ""  # 항목 칼럼이 상이할 시 col, col2, ...에서 지정하는 순서
v_manual_insert_chk = False

v_download_folder = r"V:\document\정보공시데이터\PYTHON\excel\down_file"
v_file_path = v_download_folder + "\\"  # 엑셀 다운로드 경로

# 페이지 내 클릭에 의한 다음 루트까지의 요소 검색 지연 처리
def page_flow(browser_element):
    v_count2 = 0
    v_check_btn = False
    while not v_check_btn:
        sel_element = browser_element
        if sel_element:
            try:
                sel_element[0].send_keys("\n") if isinstance(sel_element, list) else sel_element.send_keys("\n")
                v_check_btn = True
            except ElementNotInteractableException as e:
                try:
                    browser.execute_script("arguments[0].click();", sel_element[0] if isinstance(sel_element, list) else sel_element)
                    v_check_btn = True
                except ElementNotInteractableException as e:
                    sel_element[0].click() if isinstance(sel_element, list) else sel_element.click()
                    v_check_btn = True
        else:
            v_check_btn = False

        if v_count2 == 13:
            print("요소 찾기 시간 초과")
            sys.exit()

        v_count2 += 1
        time.sleep(0.33)


def main_logic(m_iaif_name):
    print ('크롤링 시작')
    # 파라미터
    global v_year
    v_year = infoClickYearly.v_down_year
    # v_dtYear = 2018  # 기준년도 : 공시에 따라 info_year 또는 info_year-1
    global v_table
    v_table = m_iaif_name
    v_table = v_table.lower()
    v_univgrp1 = ""  # 태그의 option value - 01[전문대학], 02[대학], 03[대학원]
    global univ_grp_type
    if v_table[4:5] == "5":
        univ_grp_type = "02"
    elif v_table[4:5] == "7":
        univ_grp_type = "01"
    else:
        univ_grp_type = "03"

    v_col_univ_info = [
        "학교종류"
        , "설립구분"
        , "지역"
        , "상태"
    ]

    # 공시 파일 내 변수 가져오기
    iaif_path_append()
    global v_mod
    v_mod = __import__(v_table)

    v_iaif_name = v_mod.iaif_name
    print(v_iaif_name)

    v_iaif_ord_name = None
    try:
        v_iaif_ord_name = __import__(v_table).iaif_ord_name
    except AttributeError:
        pass

    # 대학원 취업 현황의 전문대학원과 특수대학원은 따로 공시되지만 같은 공시 테이블을 사용하므로
    # 뒤에 오는 특수대학원에 공시 테이블명을 두어 다른 파일명이지만 같은 테이블에 insert 되도록 처리
    try:
        v_table = v_mod.table_name
    except AttributeError:
        pass

    try:
        v_dtyear_idx = v_mod.dtyear_idx
    except AttributeError:
        v_dtyear = int(v_year) - int(v_mod.dtyear_num)
        v_dtyear_idx = -1  # 공시를 조회한 항목에서 넣어야 할 때, 그 외는 dtyear_num값 만큼 평가년도에서 감소

    # infoClickYearly의 v_year_data_check가 True이면 테이블에 해당 연도 데이터가 있으면 지우고, insert 없으면 미공시로 판단
    if infoClickYearly.v_year_data_check:  # True
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(1) AS CNT FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) + "'")
        global result
        result = cursor.fetchall()
        cursor.close()

        if result[0][0] == 0:  # 해당 테이블에 지정 년도의 데이터가 없다면 함수 종료
            return 0

    # kwd : 공시 검색 명, schlKnd : 학교종류코드(01:전문대학,02:대학,03:대학원)
    if v_iaif_ord_name:
        # element[0].send_keys(v_iaif_org_name)
        browser.get('http://academyinfo.go.kr/search/search.do?kwd="' + v_iaif_ord_name + '"&schlKnd=' + univ_grp_type)
    else:
        # element[0].send_keys(v_iaif_name)
        browser.get('http://academyinfo.go.kr/search/search.do?kwd="' + v_iaif_name + '"&schlKnd=' + univ_grp_type)

    # element[0].send_keys(Keys.ENTER)
    time.sleep(1.5)

    # 공시 데이터의 년도 클릭
    global element_iaif
    try:
        element_iaif = browser.find_element_by_xpath(
            "//table[@class='tbl-col']/tbody[@id='targetDiv']//td[contains(text(),'" +
            v_iaif_name + "')][not(contains(text(),'PDF'))]/parent::tr//button[@data-svy_yr='" + str(v_year) + "']")
        page_flow(element_iaif)
    except:
        try:
            element_iaif = browser.find_element_by_xpath(
                "//table[@class='tbl-col']/tbody[@id='targetDiv']//td[contains(text(),'" +
                v_iaif_ord_name + "')]/parent::tr//button[@data-svy_yr='" + str(v_year) + "']")
            page_flow(element_iaif)
        except:
            print("공시 항목을 찾을 수 없습니다.")
            ws.cell(v_cell_row, v_year_index_in_file + 1).value = "미공시"
            wb.save("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")
            return 0

    time.sleep(1.5)

    if len(browser.window_handles) > 1:
        browser.switch_to.window(browser.window_handles[1])

    global bs_soup
    bs_soup = BeautifulSoup(browser.page_source, 'html.parser')

    # 공시 항목 데이터 창 open 후, 데이터 불러오기 까지 지연 시키기
    v_check = False
    v_count3 = 0
    while not v_check and v_count3 <= 540:  # 540 : 약 3분
        bs_soup = BeautifulSoup(browser.page_source, 'html.parser')
        check_elmt = bs_soup.find("div", {"id": "UbiHTMLViewer_preview_1"})
        if check_elmt:
            v_check = True

        v_count3 += 1
        time.sleep(0.33)

    # 현재 공시와 저장된 공시의 칼럼 비교 : ext_info_columns.get_col_title
    cur_title_list = get_col_title(browser)

    try:
        global v_cols_order
        if cur_title_list == v_mod.total_col:
            v_cols_order = ""
        elif cur_title_list == v_mod.total_col2:
            v_cols_order = "2"
        elif cur_title_list == v_mod.total_col3:
            v_cols_order = "3"
    except AttributeError:
        ctypes.windll.user32.MessageBoxW(0, str(v_year) + "년도 공시의 제목 열이 기존 제목 열과 다릅니다. 확인 바랍니다.", "", 4096)
        conn.close()
        sys.exit()
    # ###

    # 모든 페이지를 불러왔는지를 확인
    v_check = False
    v_count3 = 0
    while not v_check:
        page_check = browser.find_elements_by_xpath("//div[@id='UbiHTMLViewerUbiToolbarButton_TotalPageText']")[0].text
        if "+" not in page_check:
            v_check = True
        elif v_count3 >= 540:
            ctypes.windll.user32.MessageBoxW(0, str(v_year) + "모든 페이지를 불러오지 못했습니다. 확인 바랍니다.", "", 4096)
            conn.close()
            sys.exit()

        v_count3 += 1
        time.sleep(0.33)

    # 수동 파일 처리 : excel>down_file에 엑셀파일로 다운로드함. -> infoFromExcel.py에서 수동으로 업로드
    if v_manual_insert_chk:
        print('수동 항목 다운로드')
        time.sleep(0.5)
        page_flow(browser.find_elements_by_xpath("//td[@id='UbiHTMLViewerUbiToolbar_SaveButton']/input"))  # 저장버튼 클릭
        time.sleep(2)
        v_count3 = 0
        v_check = False
        this_year = datetime.today().year
        while not v_check:  # 30초
            v_save_file_list = os.listdir(v_file_path)
            if len(v_save_file_list) > 0:
                for idx, file in enumerate(v_save_file_list):
                    if str(this_year) in file and not file.endswith(".crdownload"):
                        time.sleep(0.5)
                        v_save_file_name = os.listdir(v_file_path)[idx]
                        os.rename(v_file_path + v_save_file_name, v_file_path + v_iaif_name + ".xlsx")
                        v_check = True
                        print(v_save_file_name + " saved")
                        break

            v_count3 += 1
            if v_count3 > 100:
                print(v_iaif_name + "이 다운로드 되지 않았거나 여러 파일이 존재합니다.")
                conn.close()
                wb.close()
                browser.quit()
                sys.exit()
            time.sleep(0.3)

        browser.close()
        browser.switch_to.window(browser.window_handles[0])
        return 0
        # browser.quit()
        # conn.close()
        # sys.exit()

    scroll_element = "UbiHTMLViewer_previewframe"
    scrollElem = browser.find_element(By.ID, value=scroll_element)

    docHeight = browser.execute_script("return document.scrollingElement.scrollHeight")
    scrollHeight = browser.execute_script("return document.getElementById('"+scroll_element+"').scrollHeight")

    actions = ActionChains(browser)
    actions.move_to_element(scrollElem)
    actions.click()
    actions.perform()

    curHeight = 0
    i = 1
    while curHeight < scrollHeight:
        curHeight = docHeight * i
        browser.execute_script("document.getElementById('"+scroll_element+"').scrollTop = " + str(curHeight))
        time.sleep(1.2)
        i += 1

    bs_soup2 = BeautifulSoup(browser.page_source, 'html.parser')
    myElem3 = bs_soup2.find("div", {"id": scroll_element})

    global insert_data
    insert_data = []  # table insert data
    global column_length
    column_length = 0  # insert 행의 컬럼 길이

    my_element4 = bs_soup2.find("div", {"id": "UbiHTMLViewer_previewpage_1"})
    my_element4 = my_element4.find_all("div", {"class": "UbiHTMLViewer_previewpage_1color_b_2"})

    v_univ_title_index = None
    v_minus_val = 0
    for idx, val in enumerate(my_element4):
        if idx == 0 and ("연도" in val.get_text() or "년도" in val.get_text()):
            v_minus_val = 1
        if val.get_text() in ["학교명", "학교", "대학명", '대학원명', '대학(원)명']:
            v_univ_title_index = idx - v_minus_val
            break

    # 정보공시 대학명 검사
    def is_univ_nm_check(val):
        univ_nm_str = ["대학", "학교", "기술원", "본교", "캠퍼스", "분교"]
        gbn = False
        for str in univ_nm_str:
            if str in val:
                gbn = True
        return gbn

    # 정보공시 대학원명 대학교 뒤 공백 검사

    def is_univ_nm_space_check(val):
        univ_nm_str = ["대학교", "기술원"]
        univ_txt = ""
        for str in univ_nm_str:
            if str in val:
                univ_txt = str

        string_idx = val.index(univ_txt)
        if val[string_idx+len(univ_txt)] != " ":
            return val.replace(univ_txt, univ_txt+" ")
        else:
            return val

    for child in myElem3.children:
        div_elmt = child.findAll("div", {"class": "textitem"})
        univ_data = {}
        for idx, el in enumerate(div_elmt):
            attr_style = el.attrs["style"]
            if el.parent.get("id") + "color_b_2" not in el.attrs["class"] and el.parent.get("id")+"color_b_2_0" not in el.attrs["class"]:
                # 페이지 당 칼럼 제목과 기준년도 데이터는 제외(기타)
                re_left = re.compile("left: [0-9]{1,5}px;+")  # style에서 left 추출(px제외)
                re_top = re.compile("top: [0-9]{1,5}px;+")  # style에서 top 추출(px제외)
                css_left = re_left.search(attr_style)
                css_top = re_top.search(attr_style)
                css_left = css_left.group().replace("px;", "").split(" ")[1]
                css_top = css_top.group().replace("px;", "").split(" ")[1]

                # left와 top 스타일 수치를 이용하여 dictionary 생성
                def css_grouping(v_css_left, v_css_top):
                    data = ""
                    # 대학명이면 "_" 이전 공백 제거 또는 "_"가 없는 대학교명의 맨 끝 공백 제거, 수치이면 천단위 콤마 제거
                    if is_univ_nm_check(el.get_text()):
                        data = el.get_text().strip().replace(" _", "_")
                    elif v_table in ("iaif5200_13", "iaif6200", "iaif7200_13", "iaif5341","iaif5541"):
                        # 졸업생 취업률의 취업률 칼럼은 16년 이전에는 '-'로 표시되었기 때문에 ''로 수정
                        data = "" if el.get_text() == "-" else el.get_text().replace(",", "")
                    else:
                        # 수치 데이터의 ','(comma) 표시 제거
                        data = el.get_text().replace(",", "")  # .replace("∙", "ㆍ")

                    if v_univgrp1 == "03":
                        data = is_univ_nm_space_check(el.get_text())

                    if css_top in univ_data:
                        univ_data[v_css_top][v_css_left] = data
                    else:
                        univ_data[v_css_top] = {v_css_left: data}

                if int(css_top) > 55:
                    if v_dtyear_idx > -1 or int(css_left) > 0:
                        css_grouping(css_left, css_top)

        row_list = []
        bf_list = []
        univ_data = sorted(univ_data.items(), key=lambda x: int(x[0]))

        first_element_cnt = 0  # insert 데이터의 첫번째 요소의 수
        for key, val in univ_data:
            # top(key)별로 값({left: data, ...})을 left순으로 정렬
            sort_item = sorted(val.items(), key=lambda x: int(x[0]))
            item_list = list(list(zip(*sort_item))[1])
            key_list = list(list(zip(*sort_item))[0])

            # style left 값으로 기준 list를 만들고, 이것으로 각 행을 비교하여 병합된 부분을 이전 행에서 메꾼다.
            if not row_list:
                row_list = key_list[:]
            elif len(row_list) != len(key_list):
                for idx, val in enumerate(row_list):
                    if idx == len(key_list) or val != key_list[idx]:
                        key_list.insert(idx, val)
                        item_list.insert(idx, bf_list[idx])

            bf_list = item_list[:]  # 이전 리스트를 다음 행에서 참조하기 위해 할당함

            # 대학명의 위치를 변경하는 부분 : 현장실습 현황(iaif5515_15,18)의 경우 대학명이 학교 정보보다 앞에 있음(맨 앞)
            try:
                v_univ_title_index = 4 if v_mod.univ_name_pos_idx > -1 else None
                item_list.insert(4, item_list.pop(v_mod.univ_name_pos_idx))
            except AttributeError:
                pass

            # 기준년도 칼럼 추출. (학자금 대출 현황, 취업률 현황(전문대학원, 특수대학원), 졸업생의 진학현황(전문, 특수대학원))
            v_dtyear_element = [item_list.pop(v_dtyear_idx)] if v_dtyear_idx > -1 else []

            # 대학명이 존재하는 칼럼 index 할당 : 일부 공시는 학교종류가 병합될 수 도 있음
            # 보통 데이터는 대학명부터 잘라서 집어넣는데, 대학원(신입생 충원 현황)의 경우 대학원종류 칼럼부터 insert하므로 이를 처리
            try:
                univ_nm_idx = v_mod.main_data_start_index
            except AttributeError:
                univ_nm_idx = v_univ_title_index

            # if is_univ_nm_check(item_list[5]):
            #     univ_nm_idx = 5
            # elif is_univ_nm_check(item_list[4]):
            #     univ_nm_idx = 4
            # elif is_univ_nm_check(item_list[3]):
            #     univ_nm_idx = 3
            # elif is_univ_nm_check(item_list[2]):
            #     univ_nm_idx = 2
            # elif is_univ_nm_check(item_list[1]):
            #     univ_nm_idx = 1
            # elif is_univ_nm_check(item_list[0]):
            #     univ_nm_idx = 0

            # 정렬 후 값만 추출하여 insert_data list에 추가 : 재적 학생 현황 공시에 해당
            v_univ_info = item_list[:univ_nm_idx] if v_mod.univ_info_add else []  # 학교정보 추가

            # 일부 공시는 테이블이 분리되어 있기 때문에 나눠서 추출한다. (국내외 연구실적, 저역서 실적)
            v_div_list_module, v_data_div_list = [], []
            try:
                if v_mod.data_div_rng:
                    # v_cols_order의 순서에 따라 data_div_rng의 번호를 할당
                    if v_cols_order == "":
                        v_div_list_module = v_mod.data_div_rng
                    elif v_cols_order == "2":
                        v_div_list_module = v_mod.data_div_rng2
                    elif v_cols_order == "3":
                        v_div_list_module = v_mod.data_div_rng3

                for idx, value in enumerate(v_div_list_module):
                    v_data_div_list += v_dtyear_element + (
                        item_list[univ_nm_idx:value] if idx == 0 else item_list[value[0]:value[1]])
            except AttributeError:
                v_data_div_list = None
                pass

            # 특정 공시의 경우 기준년도가 첫 열이 아닌 다른 열에 존재할 수도 있기 때문에 이를 아래와 같이 처리함
            insert_year = [str(v_year)]
            try:
                if v_mod.dtyear_idx:
                    pass
            except AttributeError:
                insert_year.append(str(v_dtyear))

            # None 데이터를 지정된 데이터로 대체하기 위함. item_list는 기준년도 칼럼 제외되어 있음. 주의 v_mod.data_modify[0]
            try:
                if v_mod.data_modify:
                    v_data_modify_list = v_mod.data_modify
                    for idx, val in enumerate(item_list):
                        if idx == v_data_modify_list[0] and val == "":
                            item_list[idx] = v_data_modify_list[1]
            except AttributeError:
                pass

            # 특정 위치의 요소를 대학명부터 추출하는 item_list에 지정된 위치로 삽입하기 위함
            try:
                if v_mod.change_element:
                    v_change_list = v_mod.change_element
                    v_insert_element = item_list[v_change_list[0]]  # change_element의 첫 번째 요소(추출 인덱스)
                    v_main_list = item_list[univ_nm_idx:]
                    v_main_list.insert(v_change_list[1], v_insert_element)  # change_element의 두 번째 요소(삽입 인덱스)
            except AttributeError:
                v_main_list = item_list[univ_nm_idx:]

            append_data = insert_year + v_dtyear_element + (v_main_list if not v_data_div_list else v_data_div_list) + v_univ_info

            try:
                merge_check = v_mod.merge_cell_list
            except AttributeError:
                merge_check = False

            if v_table == "iaif5354":  # 학비감면 준수 여부 항목 처리(iaif5354)
                # INSERT되는 데이터 순서 : 10% 금액, 30% 금액, 학부+대학원 등록금수입, 학부 등록금수입
                if len(insert_data) > 0 and insert_data[len(insert_data) - 1][2] == append_data[2]:
                    # 대학별 두번째 데이터는 30%규정준수금액만 추출하여 이전 동일 데이터에 insert
                    insert_data[len(insert_data) - 1].insert(4, append_data.pop(8))
                else:
                    append_data.append(append_data.pop(5))  # index 5 요소를 맨 끝으로 이동 : 등록금수입총액(학부) 이동
                    del append_data[3:7]  # index 3부터 7번째 요소까지 index기준 3~6 제거
                    insert_data.append(append_data)
            elif merge_check:
                if first_element_cnt == 0 or first_element_cnt == len(append_data):
                    insert_data.append(append_data)
                else:
                    for m_idx in merge_check:
                        append_data.insert(m_idx, insert_data[len(insert_data) - 1][m_idx])
                    insert_data.append(append_data)
            else:
                insert_data.append(append_data)

            if len(insert_data) == 1:
                print(insert_data[0])
                first_element_cnt = len(insert_data[0])
                column_length = len(insert_data[0])


def show_data(data):
    for val in data:
        print(val)


def insert_logic():  # 테이블 insert 처리
    # print(insert_data)
    # show_data(insert_data)
    v_table_cols = ""
    print('col ord : '+v_cols_order)
    if v_cols_order == "":
        v_table_cols = """(""" + ",".join(v_mod.table_cols) + """)""" if v_mod.table_cols else ""
    elif v_cols_order == "2":
        v_table_cols = """(""" + ",".join(v_mod.table_cols2) + """)""" if v_mod.table_cols2 else ""
    elif v_cols_order == "3":
        v_table_cols = """(""" + ",".join(v_mod.table_cols3) + """)""" if v_mod.table_cols3 else ""

    insert_sql = """INSERT INTO """ + v_table + v_table_cols + """ VALUES (""" + get_insert_text(column_length) + \
                 v_mod.insert_last_col + """)"""
    cursor_ins = conn.cursor()
    try:
        # insert_data = unicode(insert_data, "euc-kr").encode("utf-8")
        cursor_ins.executemany(insert_sql, insert_data)
        cursor_ins.close()
        conn.commit()
        print(v_table + " ------ update success!")
        ws.cell(v_cell_row, v_year_index_in_file + 1).value = "O"  # 엑셀 칼럼은 1부터 시작하므로 년도 추출 변수에 1을 더함
        wb.save("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(insert_sql)
        print("error code : " + str(error.code))
        print(error.message)
        print(error.context)
        cursor_ins.close()
        conn.rollback()


def insert_msg(value):  # 공시와 테이블 비교에 따른 구분
    if value == 1:  # 예 클릭
        try:
            if v_mod.delete_except:
                print("해당 항목은 Delete 제외 공시입니다.")
                pass
        except AttributeError:
            # 기데이터 삭제
            cursor_del = conn.cursor()
            cursor_del.execute("DELETE FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) + "'")
            time.sleep(2)
            cursor_del.close()
        time.sleep(0.33)
        insert_logic()
    else:
        browser.quit()
        conn.rollback()
        conn.close()
        sys.exit()


if infoClickYearly.v_total_insert_yn == "Y":

    chrome_option=webdriver.ChromeOptions()
    prefs = {"download.default_directory": v_download_folder}
    chrome_option.add_experimental_option("prefs", prefs)
    browser = webdriver.Chrome(options=chrome_option)
    browser.maximize_window()
    browser.get('chrome://settings/')
    browser.execute_script('chrome.settingsPrivate.setDefaultZoom(0.9);')
    time.sleep(1)

    browser_handle_quit(browser)

    # oracle 연동
    dsn = cx_Oracle.makedsn("61.81.234.137", 1521, "COGDW")
    conn = cx_Oracle.connect("dusd", "dusd$#@!", dsn)

    # 공시항목 크롤링 처리 로직
    browser.get("http://academyinfo.go.kr/index.do")
    time.sleep(1.5)
    browser_handle_quit(browser)

    soup = BeautifulSoup(browser.page_source, 'html.parser')

    # 정보공시 재업데이트 공시 목록 불러오기
    wb = pyxl.load_workbook("V:\\document\\정보공시데이터\\정보공시 재업데이트 관리_PYTHON.xlsx")
    ws = wb[v_sheet_name]  # or wb.active : 활성화된 시트
    v_year_index_in_file = 2

    # infoClickYearly 파일의 지정 년도가 재업데이트 파일에서의 칼럼의 년도에서의 위치 추출
    for idx, col in enumerate(list(ws.rows)[0]):
        if str(col.value) == str(infoClickYearly.v_down_year):
            v_year_index_in_file = idx
            break

    for r in list(ws.rows)[1:]:
        insert_chk = r[v_year_index_in_file].value
        # 엑셀파일에서 해당 년도 값이 비었으면 항목 추출 실행
        if not insert_chk:
            v_cell_row = r[0].row
            global v_file_iaif_name
            v_file_iaif_name = r[0].value
            print("------------ 구분선 -------------")
            print("공시 항목 : " + v_file_iaif_name)
            if r[1].value == "수동":
                print("해당 항목은 수동 작업 대상입니다,")
                v_manual_insert_chk = True
            else:
                v_manual_insert_chk = False

            if main_logic(v_file_iaif_name) != 0:
                browser.close()
                browser.switch_to.window(browser.window_handles[0])

                insert_msg(1)
                # diff = len(insert_data) - result[0][0]
                # if result[0][0] == 0 or diff == 0:
                #     insert_msg(1)
                # elif diff > 0:
                #     return_val = ctypes.windll.user32.MessageBoxW(0, "공시의 행의 갯수가 table보다 " + str(
                #         diff) + "개 많습니다. 계속 진행할까요?", "", 4097)
                #     insert_msg(return_val)
                # elif diff < 0:
                #     return_val = ctypes.windll.user32.MessageBoxW(0, "공시의 행의 갯수가 table보다 " + str(
                #         abs(diff)) + "개 적습니다. 계속 진행할까요?", "", 4097)
                #     insert_msg(return_val)


conn.close()
wb.close()
browser.quit()
