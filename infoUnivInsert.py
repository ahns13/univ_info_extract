# 개별대학 항목 추출 로직 IAIF_DATA_COMP_US에서 해당 테이블의 DATA가 비교해서 다른 대학만 추출
import time
import sys
import openpyxl as pyxl
import keyboard
import pyperclip
import os.path
import cx_Oracle
import ctypes

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from datetime import datetime
from info_func import *

os.environ["NLS_LANG"] = ".AL32UTF8"  # DB 케릭터셋과 일치시킴

# 변수
insert_data, culumn_length, v_cell_row, v_insert_check = [], 0, 0, False
wb, ws = "", ""
column_length = 0
new_ins_check = True  # 테이블에 데이터가 있으면 False, 신규 insert이면 True
# v_table: iaifList의 공시파일명
# v_table_org: DATA_COMP_US 테이블에 있는 대학 공시 테이블명, 학과가 존재하는 대학 공시 테이블
v_year = int(input("공시년도 : "))
# v_dtYear = 2018  # 기준년도 : 공시에 따라 info_year 또는 info_year-1

print("생성된 윈도우 창에서 테이블 목록 중 하나를 선택하세요.")
uList = __import__("infoUnivClickList")
time.sleep(0.5)
v_table = uList.check_us_name
# v_table = input("공시 테이블명 : ")  # 파이썬 파일명을 그대로 입력
v_table = v_table.lower()
# v_table_org = input("DATA_COMP 대상 테이블명 : ")  # IAIF_DATA_COMP_US의 테이블명
v_table_org = uList.check_ref_name
v_table_org = v_table_org.upper()

v_cols_order = ""  # input('칼럼 목록 순서[2, 3 or Enter] : ')  # 항목 칼럼이 상이할 시 col, col2, ...에서 지정하는 순서

v_read_only = input_func("읽기 전용 여부(Y|N)", "str", ["Y", "N"])  # 읽기 전용 시 병합 체크 로직 처리 못함
v_read_only = v_read_only.upper()

v_save_exec = input_func("저장 실행 여부(Y|N)", "str", ["Y", "N"])  # 다운로드 한 후 다시 열어서 저장을 하는 과정 실행 여부
v_save_exec = v_save_exec.upper()

v_new_only_insert = input_func("테이블에 없는 대학만 INSERT(Y|N)", "str", ["Y", "N"])  # 테이블에 있으면 PASS
v_new_only_insert = v_new_only_insert.upper()

v_file_ext = ".xlsx"

if v_table[4:5] == "5":
    univ_grp_type = "A"
elif v_table[4:5] == "7":
    univ_grp_type = "C"
else:
    univ_grp_type = "B"
v_univ_info_except = False

v_download_folder = r"V:\document\정보공시데이터\PYTHON\excel\down_file2"
v_file_path = v_download_folder + "\\"  # 엑셀 다운로드 경로

time.sleep(0.33)

v_col_univ_info = [
    "학교종류",
    "설립구분",
    "지역",
    "상태"
]

# 공시 파일 내 변수 가져오기
iaif_path_append()
v_mod = __import__(v_table)

v_total_col = v_mod.total_col
v_iaif_name = v_mod.iaif_name

print(v_table, ":", v_iaif_name)
v_iaif_path = None
try:
    v_iaif_path = __import__(v_table).iaif_path
except AttributeError:
    pass

# 대학원 취업 현황의 전문대학원과 특수대학원은 따로 공시되지만 같은 공시 테이블을 사용하므로
try:
    v_table = v_mod.table_name
except AttributeError:
    pass

try:
    v_dtyear_idx = v_mod.dtyear_idx
except AttributeError:
    v_dtyear = v_year - v_mod.dtyear_num
    v_dtyear_idx = -1  # 공시를 조회한 항목에서 넣어야 할 때, 그 외는 dtyear_num값 만큼 평가년도에서 감소

v_column_start = v_mod.column_start_col


def insert_logic(m_mod, m_univ_nm):  # 테이블 insert 처리
    # print_lists(insert_data)
    print(insert_data[:2])
    v_table_cols = ""
    if v_cols_order == "":
        v_table_cols = """(""" + ",".join(v_mod.table_cols) + """)""" if v_mod.table_cols else ""
    elif v_cols_order == "2":
        v_table_cols = """(""" + ",".join(v_mod.table_cols2) + """)""" if v_mod.table_cols2 else ""
    elif v_cols_order == "3":
        v_table_cols = """(""" + ",".join(v_mod.table_cols3) + """)""" if v_mod.table_cols3 else ""

    insert_sql = "INSERT INTO " + v_table + v_table_cols + " VALUES ('" + str(v_year) + "','" + str(v_dtyear) + "','" +\
                 m_univ_nm + "'," + get_insert_text(column_length) + "null," + m_mod.insert_last_col + ")"
    cursor_ins = conn.cursor()
    try:
        # insert_data = unicode(insert_data, "euc-kr").encode("utf-8")
        cursor_ins.executemany(insert_sql, insert_data)
        cursor_ins.close()
        conn.commit()
        print(v_table + " " + m_univ_nm + " insert success!")
        if v_table not in ("iaif5354", "iaif7354"):
            ws.cell(row=v_cell_row, column=4).value = "INSERT"
        else:
            ws.cell(row=v_cell_row, column=5).value = "INSERT"

        wb.save(file_path + v_iaif_name + v_file_add_text + ".xlsx")

        # if "수정" in v_new_file_name and os.path.isfile(v_old_file_name):
        #     os.remove(v_old_file_name)
    except cx_Oracle.DatabaseError as e:
        error, = e.args
        print(insert_sql)
        print(error.code)
        print(error.message)
        print(error.context)
        cursor_ins.close()
        conn.rollback()
        conn.close()
        wb.close()
        sys.exit()


def insert_msg(m_mod, u_nm):  # 공시와 테이블 비교에 따른 구분
    if new_ins_check:
        insert_logic(m_mod, u_nm)
    else:
        cursor_chk = conn.cursor()
        univ_data_chk = cursor_chk.execute("SELECT COUNT(1) FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) +
                                           "' AND UNIV_NM = '" + str(u_nm) + "'")
        univ_ins_data = univ_data_chk.fetchall()
        # if univ_data_chk == 0:
        #     print(u_nm, "의 데이터가 존재하지 않습니다. 신규 대학인지 또는 오류인지 확인 바랍니다.")
        #     cursor_chk.close()
        #     conn.close()
        #     wb.close()
        #     sys.exit()

        if v_new_only_insert == "Y":
            cursor_chk.close()
            if univ_ins_data[0][0] == 0:
                insert_logic(m_mod, u_nm)
            else:
                print("PASS : 데이터가 존재합니다.")
        else:
            # 기데이터 삭제
            cursor_chk.execute(
                "DELETE FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) + "' AND UNIV_NM = '" + str(u_nm) + "'")
            time.sleep(2)
            cursor_chk.close()
            time.sleep(0.33)
            insert_logic(m_mod, u_nm)


# 대학 공시 검색 --start
def univ_search(u_nm):
    global save_univ_name, v_new_file_name, v_old_file_name, insert_data
    v_old_file_name = v_file_path + u_nm + v_file_ext
    insert_data = []

    if v_save_exec == "Y":
        if v_table in ("iaif5357us", "iaif5354", "iaif7357us", "iaif7354"):
            save_univ_name = u_nm + "_수정"
            v_new_file_name = v_file_path + save_univ_name + v_file_ext

            if not os.path.isfile(v_new_file_name):
                # iaif5357us이면 파일을 열고, 신뢰(Y클릭), 읽기전용(R클릭), 닫기(C클릭)
                os.startfile(v_old_file_name)
                time.sleep(3)
                v_check = False
                v_count = 0
                while not v_check:
                    pyperclip.copy("")
                    time.sleep(0.33)
                    keyboard.press_and_release("control+c")
                    time.sleep(0.5)
                    cell_text = pyperclip.paste()
                    if str(v_year) in cell_text:
                        print(" - 정상 파일")
                        time.sleep(0.33)
                        keyboard.press_and_release("control+s")
                        time.sleep(0.5)
                        keyboard.press_and_release("alt+f+x")
                        time.sleep(1.5)
                        v_new_file_name = v_file_path + u_nm + v_file_ext
                        break

                    v_count += 1
                    if v_count > 4:
                        print(" - 오류 파일")
                        time.sleep(0.33)
                        keyboard.press_and_release("y")
                        time.sleep(2)
                        keyboard.press_and_release("r")
                        time.sleep(3)
                        keyboard.press_and_release("c")

                        time.sleep(1.75)
                        keyboard.press_and_release("alt+f+a")
                        time.sleep(2.7)
                        keyboard.press_and_release("y+3")
                        time.sleep(0.66)
                        pyperclip.copy(save_univ_name)
                        keyboard.press_and_release("control+v")
                        time.sleep(1.2)
                        keyboard.press_and_release("enter")
                        time.sleep(3.3)
                        keyboard.press_and_release("alt+f+x")
                        time.sleep(1.7)
                        break
                    time.sleep(0.3)

            time.sleep(0.33)
        else:
            v_new_file_name = v_file_path + u_nm + v_file_ext
            if os.path.isfile(v_new_file_name):
                os.startfile(v_new_file_name)
                time.sleep(3)
                v_check = False
                v_count = 0
                while not v_check:
                    pyperclip.copy("")
                    time.sleep(0.33)
                    keyboard.press_and_release("control+c")
                    time.sleep(0.5)
                    cell_text = pyperclip.paste()
                    if str(v_year) in cell_text:
                        print(" - 정상 파일")
                        time.sleep(0.33)
                        keyboard.press_and_release("control+s")
                        time.sleep(0.65)
                        keyboard.press_and_release("alt+f+x")
                        time.sleep(1.5)
                        break

                    v_count += 1
                    if v_count > 4:
                        print("엑셀 파일 열기 - 저장에 오류가 발생하였습니다.")
                        conn.close()
                        wb.close()
                        sys.exit()

                    time.sleep(0.3)
    else:
        v_new_file_name = v_file_path + u_nm + "_수정" + v_file_ext
        if not os.path.isfile(v_new_file_name):
            v_new_file_name = v_new_file_name.replace("_수정", "")

    time.sleep(0.5)
    # 다운받은 파일을 열어서 데이터 추출
    global wb_info, ws_info, column_length, row_value
    v_read_only_chk = True if v_read_only == "Y" else False
    wb_info = pyxl.load_workbook(filename=v_new_file_name, read_only=v_read_only_chk)
    ws_info = wb_info.active  # or wb.active : 활성화된 시트

    v_compare_row_num = v_mod.excel_data_start_row
    if v_table in ("iaif5354", "iaif7354"):
        # 등록금 수입 추출
        v_find_row_count = 999
        item_list = []

        ws_row_list = list(ws_info.rows)
        v_last_row_except_num = 1 if ws_row_list[len(ws_row_list) - 1][0].value is None else 0

        sust_index = get_sust_col_index(ws_row_list)

        for idx, row in enumerate(list(ws_info.rows)[v_compare_row_num:len(ws_row_list) - v_last_row_except_num]):
            if v_find_row_count < 999:
                v_find_row_count -= 1
            if len(row) > 0 and v_mod.cell_find_text in str(row[0].value):
                v_find_row_count = v_mod.cell_find_add_num
            if v_find_row_count == 0:
                if v_table == "iaif7354":
                    for d_idx, d_value in enumerate(list(row)):  # 지정된 시작 열부터 포함
                        row_value = str(d_value.value)
                        if row_value not in ["None", ""]:
                            if sust_index is None or d_idx != sust_index:
                                row_value = row_value.replace(",", "")
                                item_list.append(row_value)
                elif v_table == "iaif5354":
                    if len(item_list) == 0:
                        item_list.append(row[v_mod.ext_cell_cols1[0]].value)  # 학부+대학원 10%금액
                        item_list.append(row[v_mod.ext_cell_cols1[1]].value)  # 학부+대학원 등록금 수입 총액
                        item_list.append(row[v_mod.ext_cell_cols1[2]].value)  # 학부 등록금 수입 총액
                        v_find_row_count += 1
                        continue
                    else:
                        item_list.append(row[v_mod.ext_cell_cols2[0]].value)  # 학부+대학원 30%금액

                if len(item_list) > 0:
                    # insert_year = [str(v_year), str(v_dtyear)]
                    insert_data.append(item_list)
                else:
                    continue

                if len(insert_data) == 1:
                    # print(insert_data[0])
                    column_length = len(insert_data[0])
                break

        if v_find_row_count == 999:
            print("등록금 수입 항목 추출에 실패하였습니다.")
            conn.close()
            wb.close()
            sys.exit()
    elif v_table in ("iaif5343us", "iaif5342us"):
        if_lvl1 = v_mod.dict_if_1
        if_lvl2 = v_mod.dict_if_2
        if_lvl3 = v_mod.dict_if_3
        try:
            if_lvl4 = v_mod.dict_if_4
        except AttributeError:
            pass

        ws_row_list = list(ws_info.rows)
        v_last_row_except_num = 1 if ws_row_list[len(ws_row_list) - 1][0].value is None else 0

        sust_index = get_sust_col_index(ws_row_list)

        for idx, row in enumerate(list(ws_info.rows)[v_compare_row_num:len(ws_row_list) - v_last_row_except_num]):
            item_list = []
            row_level = 0
            for d_idx, d_value in enumerate(list(row)[v_column_start:]):  # 지정된 시작 열부터 포함
                row_value = str(d_value.value)
                global if_no
                if_no = ""
                if d_idx == 0:
                    if row_value in list(if_lvl1.keys()):  # 1레벨
                        row_level = if_lvl1[row_value][0]
                        if_no = if_lvl1[row_value][1]
                    elif "- " in row_value:  # 3레벨 이상
                        row_value = row_value.replace("- ", "")
                        try:
                            row_level = if_lvl3[row_value][0]
                            if_no = if_lvl3[row_value][1]
                        except KeyError:
                            row_level = if_lvl4[row_value][0]
                            if_no = if_lvl4[row_value][1]
                    else:
                        row_level = if_lvl2[row_value][0]
                        if_no = if_lvl2[row_value][1]
                    item_list.append(if_no)
                else:
                    if sust_index is None or d_idx != sust_index - v_column_start:
                        row_value = row_value.replace(",", "")
                if row_level > 0:
                    item_list.append(row_value)

            if len(item_list) > 0:
                # insert_year = [str(v_year), str(v_dtyear)]
                insert_data.append(item_list)
            else:
                continue

            if len(insert_data) == 1:
                # print(insert_data[0])
                column_length = len(insert_data[0])
    else:
        ws_row_list = list(ws_info.rows)
        v_last_row_except_num = 1 if ws_row_list[len(ws_row_list) - 1][0].value is None else 0
        sust_index = get_sust_col_index(ws_row_list)
        for idx, row in enumerate(list(ws_info.rows)[v_compare_row_num:len(ws_row_list) - v_last_row_except_num]):
            item_list = []
            # 행 제외 데이터에 포함되면 해당 행 부분은 insert하지 않고 다음 반복으로 이동
            # us 파일에 delete_row_compare_col_index 변수 선언
            try:
                if (str(row[v_mod.delete_row_compare_col_index[0]].value) in v_mod.delete_row_compare_col_index[1]
                   or row[v_mod.delete_row_compare_col_index[0]].value is None):
                    continue
            except AttributeError:
                pass
            except IndexError:  # 행의 데이터 수가 부족한 행은 일반적인 데이터 행이 아니므로 제외
                continue

            # 불필요한 열 병합으로 인한 빈 칼럼에 해당하는 데이터 제외하기
            # us 파일에 delete_col_index_list 변수 선언
            try:
                row = [el for no, el in enumerate(row) if no not in v_mod.delete_col_index_list]
            except AttributeError:
                pass

            for d_idx, d_value in enumerate(list(row)[v_column_start:]):  # 지정된 시작 열부터 포함
                row_value = str(d_value.value)
                try:
                    if d_idx in v_mod.merge_check_cols:
                        # 데이터 행에 대한 시작 열을 지정해서 가져오기 때문에 그 수치만큼 더하여야 한다.(univ_nm_idx)
                        if row_value in ["None", ""]:
                            row_value = insert_data[len(insert_data) - 1][d_idx]  # 이전 행의 해당 인덱스의 요소 할당
                except AttributeError:
                    pass

                try:
                    if d_idx != v_mod.sust_comma_except_col_no:
                        if sust_index is None or d_idx != sust_index - v_column_start:
                            row_value = row_value.replace(",", "")
                            row_value = "" if row_value == "None" else row_value
                except AttributeError:
                    if sust_index is None or d_idx != sust_index - v_column_start:
                        row_value = row_value.replace(",", "")
                        row_value = "" if row_value == "None" else row_value

                try:
                    if d_idx == v_mod.data_change[0]:
                        row_value = v_mod.data_change[2] if row_value == v_mod.data_change[1] else row_value
                except AttributeError:
                    pass

                item_list.append(row_value)

            if len(item_list) > 0:
                # insert_year = [str(v_year), str(v_dtyear)]
                insert_data.append(item_list)
            else:
                continue

            if len(insert_data) == 1:
                # print(insert_data[0])
                column_length = len(insert_data[0])

    wb_info.close()
    if len(insert_data) > 0:
        print("데이터 수 :", len(insert_data))
        insert_msg(v_mod, u_nm)
    else:
        print("insert_data에 추출된 데이터가 없습니다.")
        ws.cell(row=v_cell_row, column=4).value = "데이터 없음"
        wb.save(file_path + v_iaif_name + v_file_add_text + ".xlsx")
        return

    time.sleep(0.33)
# 대학 공시 검색 --end


# oracle 연동
dsn = cx_Oracle.makedsn("61.81.234.137", 1521, "COGDW")
conn = cx_Oracle.connect("dusd", "dusd$#@!", dsn)
cursor = conn.cursor()
cursor.execute("SELECT COUNT(1) FROM " + v_table + " WHERE INFO_YYYY ='" + str(v_year) + "'")
result = cursor.fetchall()
new_ins_check = True if result[0][0] == 0 else False
cursor.close()

v_file_add_text = "_전문대학" if univ_grp_type == "C" else ""
file_path = r"V:\document\정보공시데이터\PYTHON\excel\\"
wb = pyxl.load_workbook(file_path+v_iaif_name+v_file_add_text+".xlsx")
ws = wb['Sheet1']  # or .active

for r in list(ws.rows)[1:]:
    v_cell_row = r[0].row
    print(r[4].value)
    v_insert_check = True
    if v_table in ("iaif5354", "iaif7354"):
        if r[3].value in ("O", "INSERT") and r[4].value is None:
            pass
        else:
            v_insert_check = False
    elif r[3].value != "O":
        v_insert_check = False

    if v_insert_check:
        print(r[0].value, ": start")
        univ_search(r[0].value)
        print('--------------------------')

print(insert_data[:3])
conn.close()
wb.close()
sys.exit()
